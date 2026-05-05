# -*- coding: utf-8 -*-
"""
phase1_dry_run.py
End-to-end Phase 1 EM-import dry-run against the real HR Excel.

Mirrors the JS pipeline:
  1. Read Excel → cleanDataset-equivalent normalization
  2. buildEmployeeMasterImportPreview() with EMPTY existing stores
     (simulates first-time import into a fresh DB)
  3. Compute the would-commit log:
        - persons (NEW)
        - employeeMasterSnapshots (NEW)
        - employeeNumberHistory (entries per person with non-empty EmpNo)
        - importAuditLog (creation entries — one per new person)
        - reviewQueue (missing + invalid + needs-review items)
        - importJobs (1 record)
  4. NO DB writes — would-commit log is exported as Excel/JSON.

Mirrors logic verified line-by-line against:
  src/services/imports/employeeMasterImportService.js
  src/services/imports/employeeNumberHistoryService.js
  src/services/imports/importCommitService.js
  src/utils/identityModel.js
  src/utils/cleaning.js
"""

import json, re, sys, datetime as dt
from pathlib import Path
from collections import Counter

import openpyxl
import pandas as pd

ROOT       = Path(__file__).resolve().parents[2]
EM_EXCEL   = Path(__file__).parent / "inputs" / "بيانات الموظفين.xlsx"
OUT_DIR    = ROOT / "_contract_lab" / "outputs"
OUT_PREV_X = OUT_DIR / "phase1_em_import_preview.xlsx"
OUT_PREV_J = OUT_DIR / "phase1_em_import_preview.json"
OUT_COMMIT = OUT_DIR / "phase1_em_would_commit.xlsx"
OUT_REPORT = OUT_DIR / "phase1_em_dry_run_report.txt"

# Mirrors schema.js schemaAliasesArabic (the relevant subset for this file)
ARABIC_ALIASES = {
    'الرقم الوطني':       'IdentityNumber',
    'رمز الموظف':         'EmployeeNumber',
    'اسم الموظف':         'Name',
    'إجمالي الراتب':      'GrossCashMonthly',
    'تاريخ بدء العقد':    'StartDate',
    'تاريخ نهاية العقد':  'EndDate',
    'الجنسية':            'Nationality',
    'الموقع':             'Location',
    'نوع العقد':          'ContractType',
    'المسمى الوظيفي':     'Profession',
    'تاريخ التعيين':      'JoiningDate',
    'تاريخ الولادة':      'DateOfBirth',
    'الجنس':              'Gender',
    'العمر':              'Age',
    'مدة الخدمة':         'ServiceDuration',
    'التأمين الصحي':      'HealthInsuranceStatus',
}

SNAPSHOT_TO_ROW_FIELDS = {
    'employeeNumber':        'EmployeeNumber',
    'sourceFile':            'SourceFile',
    'location':              'Location',
    'profession':             'Profession',
    'grossSalary':           'GrossCashMonthly',
    'healthInsuranceStatus': 'HealthInsuranceStatus',
    'contractType':          'ContractType',
    'startDate':             'StartDate',
    'endDate':               'EndDate',
    'joiningDate':            'JoiningDate',
    'dateOfBirth':           'DateOfBirth',
}


# ── helpers (mirrors cleaning.js / identityModel.js) ──────────────────────────

def norm_id(value):
    if value is None: return ''
    if isinstance(value, float):
        return '' if value != value else str(int(round(value)))
    if isinstance(value, int):
        return str(value)
    return re.sub(r'[^0-9]', '', str(value).strip())


def validate_id(value):
    id_ = norm_id(value)
    if not id_:        return {'valid': False, 'type': None, 'reason': 'missing'}
    if len(id_) != 10: return {'valid': False, 'type': None, 'reason': f'length is {len(id_)}, expected 10'}
    if id_[0] == '1':  return {'valid': True,  'type': 'Saudi', 'reason': None}
    if id_[0] == '2':  return {'valid': True,  'type': 'Iqama', 'reason': None}
    return {'valid': False, 'type': None,
            'reason': f"starts with {id_[0]}, expected 1 (Saudi) or 2 (Iqama)"}


def norm_date(value):
    if value is None: return ''
    if isinstance(value, (dt.datetime, dt.date)):
        return value.strftime('%Y-%m-%d')
    s = str(value).strip()
    if not s or s in ('None', 'null'): return ''
    m = re.match(r'^(\d{4}-\d{2}-\d{2})\s+\d{2}:\d{2}:\d{2}', s)
    if m: return m.group(1)
    if re.match(r'^\d{4}-\d{2}-\d{2}$', s): return s
    m = re.match(r'^(\d{2})-(\d{2})-(\d{4})$', s)
    if m: return f'{m.group(3)}-{m.group(2)}-{m.group(1)}'
    return s


def norm_num(value):
    if value is None or str(value).strip() in ('', 'None', 'null'): return None
    try: return float(str(value).replace(',', '').strip())
    except (ValueError, TypeError): return None


# ── reader ───────────────────────────────────────────────────────────────────

def canonical_column(col):
    raw = str(col or '').strip()
    if raw in ARABIC_ALIASES:
        return ARABIC_ALIASES[raw]
    norm = re.sub(r'[^a-z0-9]', '', raw.lower())
    en_map = {
        'identitynumber': 'IdentityNumber', 'employeenumber': 'EmployeeNumber',
        'name': 'Name', 'sourcefile': 'SourceFile',
        'startdate': 'StartDate', 'enddate': 'EndDate',
        'joiningdate': 'JoiningDate', 'dateofbirth': 'DateOfBirth',
        'grosscashmonthly': 'GrossCashMonthly', 'profession': 'Profession',
        'nationality': 'Nationality', 'gender': 'Gender',
        'location': 'Location', 'contracttype': 'ContractType',
        'age': 'Age', 'serviceduration': 'ServiceDuration',
        'healthinsurancestatus': 'HealthInsuranceStatus',
    }
    return en_map.get(norm, raw)


def read_excel(path):
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows_iter = ws.iter_rows(values_only=True)
    header_row = next(rows_iter, [])
    raw_headers = [str(c or '').strip() for c in header_row]
    mapped = [canonical_column(h) for h in raw_headers]

    rows = []
    for i, data_row in enumerate(rows_iter):
        d = {'_rowIndex': i + 2, 'SourceFile': path.name}
        for col_i, val in enumerate(data_row):
            if col_i < len(mapped):
                d[mapped[col_i]] = val
        rows.append(d)
    wb.close()
    return raw_headers, mapped, rows


def clean_dataset(rows):
    """Mirrors the relevant parts of cleanDataset() — produces the row shape
    the import service consumes."""
    cleaned = []
    for r in rows:
        c = dict(r)
        c['IdentityNumber']   = norm_id(c.get('IdentityNumber'))
        c['EmployeeNumber']   = str(c.get('EmployeeNumber') or '').strip()
        c['Name']             = str(c.get('Name') or '').strip()
        c['Nationality']      = str(c.get('Nationality') or '').strip()
        c['Location']         = str(c.get('Location') or '').strip()
        c['Profession']       = str(c.get('Profession') or '').strip()
        c['ContractType']     = str(c.get('ContractType') or '').strip()
        c['HealthInsuranceStatus'] = str(c.get('HealthInsuranceStatus') or '').strip()
        c['SourceFile']       = str(c.get('SourceFile') or '').strip()
        c['StartDate']        = norm_date(c.get('StartDate'))
        c['EndDate']          = norm_date(c.get('EndDate'))
        c['JoiningDate']      = norm_date(c.get('JoiningDate'))
        c['DateOfBirth']      = norm_date(c.get('DateOfBirth'))
        gs = norm_num(c.get('GrossCashMonthly'))
        c['GrossCashMonthly'] = gs if gs is not None else ''
        cleaned.append(c)
    return cleaned


# ── preview (mirrors buildEmployeeMasterImportPreview) ───────────────────────

def snapshot_from_row(row, import_job_id, import_date):
    out = {'identityNumber': row['IdentityNumber']}
    for snap_key, row_key in SNAPSHOT_TO_ROW_FIELDS.items():
        v = row.get(row_key)
        out[snap_key] = None if v is None or v == '' else v
    out['importJobId'] = import_job_id
    out['importDate']  = import_date
    return out


def build_preview(cleaned_rows, source_file, import_job_id):
    """Empty existing-stores scenario (first-time import)."""
    import_date = dt.datetime.now().isoformat()
    out = {
        'importJobId': import_job_id,
        'sourceType': 'EmployeeMasterExcel',
        'sourceFile': source_file,
        'generatedAt': import_date,
        'summary': {
            'total': 0, 'new': 0, 'updated': 0, 'unchanged': 0,
            'empNoHistoryCandidates': 0, 'needsReview': 0,
            'invalidIdentity': 0, 'missingIdentity': 0,
        },
        'newPersons': [],
        'updatedSnapshots': [],
        'unchangedSnapshots': [],
        'empNoHistoryCandidates': [],
        'needsReview': [],
        'invalidIdentity': [],
        'missingIdentity': [],
        'auditEntries': [],
    }

    # Detect duplicates within file
    id_count = Counter()
    for r in cleaned_rows:
        nid = r.get('IdentityNumber')
        if nid: id_count[nid] += 1

    for row in cleaned_rows:
        out['summary']['total'] += 1
        id_ = row.get('IdentityNumber') or ''
        if not id_:
            out['missingIdentity'].append({'row': row, 'reason': 'IdentityNumber blank'})
            out['summary']['missingIdentity'] += 1
            continue
        v = validate_id(id_)
        if not v['valid']:
            out['invalidIdentity'].append({'row': row, 'reason': v['reason']})
            out['summary']['invalidIdentity'] += 1
            continue
        if id_count[id_] > 1:
            out['needsReview'].append({
                'row': row,
                'reason': f"IdentityNumber {id_} appears {id_count[id_]}x in this file",
            })
            out['summary']['needsReview'] += 1
            continue

        # NEW person (no existing data)
        snap = snapshot_from_row({**row, 'IdentityNumber': id_}, import_job_id, import_date)
        person = {
            'identityNumber': id_,
            'idType':         v['type'],
            'currentName':    row.get('Name') or '',
            'nationality':    row.get('Nationality') or '',
        }
        out['newPersons'].append({'person': person, 'snapshot': snap, 'sourceRow': row})
        out['summary']['new'] += 1

        emp = (row.get('EmployeeNumber') or '').strip()
        if emp:
            hist_entry = {
                'identityNumber': id_,
                'employeeNumber': emp,
                'sourceType':     'EmployeeMasterExcel',
                'sourceFile':     row.get('SourceFile') or source_file,
                'contractNumber': '',
                'firstSeenDate':  row.get('JoiningDate') or row.get('StartDate') or import_date,
                'lastSeenDate':   row.get('EndDate') or '',
                'status':         'Active',
                'note':           '',
                'importJobId':    import_job_id,
            }
            out['empNoHistoryCandidates'].append({
                'identityNumber': id_,
                'oldEmpNo':       '',
                'newEmpNo':       emp,
                'entry':          hist_entry,
                'note':           '',
                'firstSeen':      True,
            })
            out['summary']['empNoHistoryCandidates'] += 1
    return out


# ── would-commit log (mirrors importCommitService.js) ────────────────────────

def build_would_commit(preview, imported_by=None):
    """Returns counts and lists of records that WOULD be written."""
    persons      = []
    snapshots    = []
    history      = []
    audit        = []
    review       = []

    for np in preview['newPersons']:
        persons.append(np['person'])
        snapshots.append(np['snapshot'])
        # creation audit (one per new person, matches commit service behaviour)
        audit.append({
            'action':         'create',
            'entityType':     'Person',
            'entityId':       np['person']['identityNumber'],
            'identityNumber': np['person']['identityNumber'],
            'field':          None,
            'oldValue':       None,
            'newValue': {
                'idType':         np['person']['idType'],
                'currentName':    np['person']['currentName'],
                'employeeNumber': np['snapshot'].get('employeeNumber'),
            },
            'sourceFile':    np['snapshot'].get('sourceFile') or '',
            'sourceType':    'EmployeeMasterExcel',
            'importJobId':   preview['importJobId'],
            'importedBy':    imported_by,
            'note':          'Person + EmployeeMasterSnapshot created from EM import',
        })

    # snapshot-update audit entries (none on first import, but the field exists)
    for ae in preview['auditEntries']:
        audit.append(ae)

    for h in preview['empNoHistoryCandidates']:
        history.append({**h['entry'], 'note': h.get('note') or ''})

    # review queue items
    for r in preview['missingIdentity']:
        review.append({
            'reviewType': 'MissingIdentity', 'priority': 'CRITICAL',
            'identityNumber': '', 'reason': r['reason'],
            'sourceFile': r['row'].get('SourceFile'),
            'rowIndex': r['row'].get('_rowIndex'),
        })
    for r in preview['invalidIdentity']:
        review.append({
            'reviewType': 'InvalidIdentity', 'priority': 'CRITICAL',
            'identityNumber': r['row'].get('IdentityNumber') or '',
            'reason': r['reason'],
            'sourceFile': r['row'].get('SourceFile'),
            'rowIndex': r['row'].get('_rowIndex'),
        })
    for r in preview['needsReview']:
        review.append({
            'reviewType': 'AmbiguousMatch', 'priority': 'HIGH',
            'identityNumber': r['row'].get('IdentityNumber') or '',
            'reason': r['reason'],
            'sourceFile': r['row'].get('SourceFile'),
            'rowIndex': r['row'].get('_rowIndex'),
        })

    job = {
        'id':           preview['importJobId'],
        'type':         'EmployeeMasterExcel',
        'status':       'completed',
        'sourceName':   preview['sourceFile'],
        'totalItems':   preview['summary']['total'],
        'counts': {
            **preview['summary'],
            'personsWritten':   len(persons),
            'snapshotsWritten': len(snapshots),
            'historyWritten':   len(history),
            'auditWritten':     len(audit),
            'reviewWritten':    len(review),
        },
    }

    return {
        'job':       job,
        'persons':   persons,
        'snapshots': snapshots,
        'history':   history,
        'audit':     audit,
        'review':    review,
    }


# ── reporting ────────────────────────────────────────────────────────────────

def build_report(preview, would_commit, raw_headers, mapped_headers):
    L = []

    def h(t):
        L.append(f"\n{'='*72}")
        L.append(f"  {t}")
        L.append(f"{'='*72}")

    def r(label, value):
        L.append(f"  {label:<58} {value}")

    h("PHASE 1 — END-TO-END DRY RUN (Employee Master Excel)")
    r("Input file:", EM_EXCEL.name)
    r("Source rows:", str(preview['summary']['total']))
    r("Existing-store scenario:", "EMPTY (first-time import)")
    r("Import job ID (preview):", preview['importJobId'])

    h("1. COLUMN DETECTION")
    L.append(f"  {'Raw':<30} {'Mapped'}")
    L.append(f"  {'─'*55}")
    for raw, mapped in zip(raw_headers, mapped_headers):
        L.append(f"  {raw:<30} {mapped}")

    h("2. PREVIEW SUMMARY")
    s = preview['summary']
    r("Total rows:", str(s['total']))
    r("New persons:", str(s['new']))
    r("Updated snapshots:", str(s['updated']))
    r("Unchanged snapshots:", str(s['unchanged']))
    r("EmpNo history candidates:", str(s['empNoHistoryCandidates']))
    r("Needs review:", str(s['needsReview']))
    r("Invalid identity:", str(s['invalidIdentity']))
    r("Missing identity:", str(s['missingIdentity']))

    h("3. WOULD-COMMIT LOG")
    j = would_commit['job']
    r("Stores that would be written:", "persons, employeeMasterSnapshots,")
    L.append(f"  {'':<58} employeeNumberHistory, importJobs,")
    L.append(f"  {'':<58} importAuditLog, reviewQueue")
    L.append("")
    r("persons (NEW):",                     str(len(would_commit['persons'])))
    r("employeeMasterSnapshots (NEW):",     str(len(would_commit['snapshots'])))
    r("employeeNumberHistory (NEW):",       str(len(would_commit['history'])))
    r("importAuditLog entries (NEW):",      str(len(would_commit['audit'])))
    r("reviewQueue items (NEW):",           str(len(would_commit['review'])))
    r("importJobs records (NEW):",          "1")

    h("4. PERSON BREAKDOWN")
    saudi = sum(1 for p in would_commit['persons'] if p['idType'] == 'Saudi')
    iqama = sum(1 for p in would_commit['persons'] if p['idType'] == 'Iqama')
    r("Saudi National ID (starts with 1):", str(saudi))
    r("Iqama (starts with 2):",             str(iqama))

    h("5. REVIEW QUEUE BREAKDOWN")
    rev_by_type = Counter(r['reviewType'] for r in would_commit['review'])
    rev_by_pri  = Counter(r['priority']   for r in would_commit['review'])
    for t, n in sorted(rev_by_type.items()):
        r(f"  type {t}:", str(n))
    for p, n in sorted(rev_by_pri.items()):
        r(f"  priority {p}:", str(n))

    h("6. IMPORT JOB")
    r("Job ID:",          j['id'])
    r("Type:",            j['type'])
    r("Status:",          j['status'])
    r("Source name:",     j['sourceName'])
    r("Counts:",          json.dumps(j['counts'], ensure_ascii=False))

    return "\n".join(L)


# ── main ─────────────────────────────────────────────────────────────────────

def safe_print(s):
    try: print(s)
    except UnicodeEncodeError: print(s.encode('ascii', 'replace').decode('ascii'))


def main():
    if not EM_EXCEL.exists():
        safe_print(f"ERROR: input not found: {EM_EXCEL}")
        sys.exit(1)
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    safe_print(f"Reading: {EM_EXCEL}")
    raw_headers, mapped_headers, rows = read_excel(EM_EXCEL)
    safe_print(f"  → {len(rows)} rows, {len(raw_headers)} columns")

    cleaned = clean_dataset(rows)
    import_job_id = f"phase1-dry-run-{dt.datetime.now().strftime('%Y%m%dT%H%M%S')}"

    safe_print("Building import preview (existing stores = empty)...")
    preview = build_preview(cleaned, EM_EXCEL.name, import_job_id)

    safe_print("Computing would-commit log...")
    would_commit = build_would_commit(preview)

    # JSON
    json_payload = {
        'preview': {
            **{k: v for k, v in preview.items() if k != 'auditEntries'},
            'auditEntriesCount': len(preview.get('auditEntries', [])),
        },
        'wouldCommit': would_commit,
    }
    with open(OUT_PREV_J, 'w', encoding='utf-8') as f:
        json.dump(json_payload, f, ensure_ascii=False, indent=2, default=str)
    safe_print(f"JSON:    {OUT_PREV_J}")

    # Excel — preview
    rows_for_xlsx = []
    for np in preview['newPersons']:
        rows_for_xlsx.append({
            'category':       'NEW',
            'identityNumber': np['person']['identityNumber'],
            'idType':         np['person']['idType'],
            'currentName':    np['person']['currentName'],
            'employeeNumber': np['snapshot'].get('employeeNumber') or '',
            'grossSalary':    np['snapshot'].get('grossSalary'),
            'startDate':      np['snapshot'].get('startDate') or '',
            'endDate':        np['snapshot'].get('endDate') or '',
            'location':       np['snapshot'].get('location') or '',
        })
    for r in preview['missingIdentity']:
        rows_for_xlsx.append({
            'category': 'MISSING_IDENTITY', 'rowIndex': r['row'].get('_rowIndex'),
            'currentName': r['row'].get('Name') or '',
            'employeeNumber': r['row'].get('EmployeeNumber') or '',
            'reason': r['reason'],
        })
    for r in preview['invalidIdentity']:
        rows_for_xlsx.append({
            'category': 'INVALID_IDENTITY', 'rowIndex': r['row'].get('_rowIndex'),
            'identityNumber': r['row'].get('IdentityNumber') or '',
            'currentName': r['row'].get('Name') or '',
            'reason': r['reason'],
        })
    pd.DataFrame(rows_for_xlsx).to_excel(OUT_PREV_X, index=False, engine='openpyxl')
    safe_print(f"Excel:   {OUT_PREV_X}")

    # Excel — would-commit log
    with pd.ExcelWriter(OUT_COMMIT, engine='openpyxl') as writer:
        pd.DataFrame(would_commit['persons']).to_excel(writer, sheet_name='persons', index=False)
        pd.DataFrame(would_commit['snapshots']).to_excel(writer, sheet_name='employeeMasterSnapshots', index=False)
        pd.DataFrame(would_commit['history']).to_excel(writer, sheet_name='employeeNumberHistory', index=False)
        pd.DataFrame([{
            **a,
            'newValue': json.dumps(a.get('newValue'), ensure_ascii=False, default=str),
            'oldValue': json.dumps(a.get('oldValue'), ensure_ascii=False, default=str),
        } for a in would_commit['audit']]).to_excel(writer, sheet_name='importAuditLog', index=False)
        pd.DataFrame(would_commit['review']).to_excel(writer, sheet_name='reviewQueue', index=False)
        pd.DataFrame([{
            **would_commit['job'],
            'counts': json.dumps(would_commit['job']['counts'], ensure_ascii=False),
        }]).to_excel(writer, sheet_name='importJobs', index=False)
    safe_print(f"Commit:  {OUT_COMMIT}")

    # Report
    report = build_report(preview, would_commit, raw_headers, mapped_headers)
    OUT_REPORT.write_text(report, encoding='utf-8')
    safe_print(report)
    safe_print(f"\nReport:  {OUT_REPORT}")


if __name__ == '__main__':
    main()
