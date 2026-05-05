# -*- coding: utf-8 -*-
"""
dry_run_mapped.py
Dry-run with updated Arabic column aliases — acceptance test for mapping completeness.

Input:   _contract_lab/employee_master_import_test/inputs/بيانات الموظفين.xlsx
Outputs: _contract_lab/outputs/employee_master_real_import_mapped_preview.xlsx
         _contract_lab/outputs/employee_master_real_import_mapped_preview.json
         _contract_lab/outputs/employee_master_real_import_mapped_report.txt

Acceptance condition: 0 unmapped columns for the real HR file.
"""

import sys, json, re
from pathlib import Path
from collections import Counter

import openpyxl
import pandas as pd

ROOT       = Path(__file__).resolve().parents[2]
INPUT_FILE = Path(__file__).parent / "inputs" / "بيانات الموظفين.xlsx"
OUT_DIR    = ROOT / "_contract_lab" / "outputs"
OUT_XLSX   = OUT_DIR / "employee_master_real_import_mapped_preview.xlsx"
OUT_JSON   = OUT_DIR / "employee_master_real_import_mapped_preview.json"
OUT_REPORT = OUT_DIR / "employee_master_real_import_mapped_report.txt"

# ── Arabic column aliases — mirrors schema.js schemaAliasesArabic (full updated set) ──
ARABIC_ALIASES = {
    # IdentityNumber (primary key)
    'الرقم الوطني':         'IdentityNumber',
    'رقم الهوية الوطنية':   'IdentityNumber',
    'رقم الهوية':           'IdentityNumber',
    'رقم الإقامة':          'IdentityNumber',
    'رقم الاقامة':          'IdentityNumber',
    # Name
    'الاسم':                 'Name',
    'اسم الموظف':            'Name',
    'الاسم الكامل':          'Name',
    # EmployeeNumber (secondary only — never primary match key)
    'رقم الموظف':            'EmployeeNumber',
    'الرقم الوظيفي':         'EmployeeNumber',
    'رمز الموظف':            'EmployeeNumber',    # Qiwa/SAP: "رمز" = code
    # Nationality
    'الجنسية':               'Nationality',
    # Profession
    'المسمى الوظيفي':        'Profession',
    'المهنة':                'Profession',
    # Dates
    'تاريخ الميلاد':         'DateOfBirth',
    'تاريخ الولادة':         'DateOfBirth',       # Qiwa/HR export variant
    'تاريخ بداية العقد':     'StartDate',
    'تاريخ بدء العقد':       'StartDate',         # Qiwa/HR export variant
    'تاريخ المباشرة':        'StartDate',
    'تاريخ نهاية العقد':     'EndDate',
    'تاريخ انتهاء العقد':    'EndDate',
    'تاريخ الانضمام':        'JoiningDate',
    'تاريخ التعيين':         'JoiningDate',       # appointment date = joining date
    'تاريخ انتهاء الهوية':   'IDExpiryDate',
    # Salary
    'الراتب الأساسي':        'BasicSalary',
    'الراتب':                'BasicSalary',
    'بدل السكن':             'HousingAllowance',
    'بدل النقل':             'TransportationAllowance',
    'بدل الغذاء':            'FoodAllowance',
    'إجمالي البدلات':        'TotalCashAllowances',
    'الراتب الإجمالي':       'GrossCashMonthly',
    'إجمالي الراتب':         'GrossCashMonthly',  # Qiwa/HR export variant
    # ID fields
    'نوع الهوية':            'IDType',
    # Contact
    'رقم الجوال':            'MobileNumber',
    'رقم الهاتف':            'MobileNumber',
    'البريد الإلكتروني':     'Email',
    # Banking
    'رقم الآيبان':           'IBAN',
    'الآيبان':               'IBAN',
    'اسم البنك':             'BankName',
    # Personal / demographic
    'الجنس':                 'Gender',
    'الديانة':               'Religion',
    'الحالة الاجتماعية':     'MaritalStatus',
    'المؤهل العلمي':         'Education',
    'التخصص':               'Speciality',
    'العمر':                 'Age',
    # HR-export-specific fields
    'الموقع':                'Location',
    'مدة الخدمة':            'ServiceDuration',
    'التأمين الصحي':         'HealthInsuranceStatus',
    'نوع العقد':             'ContractType',
}

_EN_RAW = {
    re.sub(r'[^a-z0-9]', '', k.lower()): v
    for k, v in {
        'SourceFile': 'SourceFile', 'ContractNumber': 'ContractNumber',
        'Name': 'Name', 'EmployeeNumber': 'EmployeeNumber',
        'Nationality': 'Nationality', 'DateOfBirth': 'DateOfBirth', 'DOB': 'DateOfBirth',
        'IdentityNumber': 'IdentityNumber', 'IDType': 'IDType',
        'IDExpiryDate': 'IDExpiryDate', 'Gender': 'Gender', 'Religion': 'Religion',
        'MaritalStatus': 'MaritalStatus', 'Education': 'Education',
        'Speciality': 'Speciality', 'IBAN': 'IBAN', 'BankName': 'BankName',
        'Email': 'Email', 'MobileNumber': 'MobileNumber',
        'ContractDurationYears': 'ContractDurationYears',
        'StartDate': 'StartDate', 'EndDate': 'EndDate', 'JoiningDate': 'JoiningDate',
        'BasicSalary': 'BasicSalary', 'HousingProvided': 'HousingProvided',
        'TransportProvided': 'TransportProvided', 'HousingAllowance': 'HousingAllowance',
        'TransportationAllowance': 'TransportationAllowance',
        'FoodAllowance': 'FoodAllowance', 'OTAllowance': 'OTAllowance',
        'MastersDegreeAllowance': 'MastersDegreeAllowance',
        'TotalCashAllowances': 'TotalCashAllowances', 'GrossCashMonthly': 'GrossCashMonthly',
        'Profession': 'Profession', 'ContractType': 'ContractType',
        'Location': 'Location', 'ServiceDuration': 'ServiceDuration',
        'HealthInsuranceStatus': 'HealthInsuranceStatus', 'Age': 'Age',
    }.items()
}


def canonical_column(col):
    raw = str(col or '').strip()
    if raw in ARABIC_ALIASES:
        return ARABIC_ALIASES[raw], 'arabic'
    norm = re.sub(r'[^a-z0-9]', '', raw.lower())
    if norm in _EN_RAW:
        return _EN_RAW[norm], 'english'
    return raw, 'unmapped'


def normalize_identity(value):
    if value is None:
        return ''
    if isinstance(value, float):
        return '' if value != value else str(int(round(value)))
    if isinstance(value, int):
        return str(value)
    return re.sub(r'[^0-9]', '', str(value).strip())


def validate_identity(value):
    id_ = normalize_identity(value)
    if not id_:
        return {'valid': False, 'type': None, 'reason': 'missing'}
    if len(id_) != 10:
        return {'valid': False, 'type': None, 'reason': f'length {len(id_)} (expected 10)'}
    if id_[0] == '1':
        return {'valid': True,  'type': 'Saudi', 'reason': None}
    if id_[0] == '2':
        return {'valid': True,  'type': 'Iqama', 'reason': None}
    return {'valid': False, 'type': None,
            'reason': f"starts with '{id_[0]}' (expected 1=Saudi or 2=Iqama)"}


def read_excel(path):
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    sheet_name = wb.sheetnames[0]
    ws = wb[sheet_name]
    rows_iter = ws.iter_rows(values_only=True)
    header_row = next(rows_iter, None)
    if header_row is None:
        wb.close()
        return sheet_name, [], [], []

    raw_headers    = [str(c or '').strip() for c in header_row]
    mapping_detail = []
    mapped_headers = []
    for raw in raw_headers:
        canonical, kind = canonical_column(raw)
        mapped_headers.append(canonical)
        mapping_detail.append((raw, canonical, kind))

    rows = []
    for data_row in rows_iter:
        d = {}
        for col_i, cell_val in enumerate(data_row):
            if col_i >= len(mapped_headers):
                break
            d[mapped_headers[col_i]] = cell_val
        rows.append(d)
    wb.close()
    return sheet_name, raw_headers, mapping_detail, rows


def analyze(rows):
    id_counter = Counter()
    results    = []
    for i, row in enumerate(rows):
        raw_id  = row.get('IdentityNumber', '')
        norm_id = normalize_identity(raw_id)
        vcheck  = validate_identity(raw_id)
        if norm_id:
            id_counter[norm_id] += 1
        results.append({
            'RowIndex':         i + 2,
            'RawIdentityValue': raw_id,
            'NormalizedID':     norm_id,
            'IDValid':          vcheck['valid'],
            'IDType':           vcheck['type'] or '',
            'IDReason':         vcheck['reason'] or '',
            'EmployeeNumber':   str(row.get('EmployeeNumber') or '').strip(),
            'Name':             str(row.get('Name') or '').strip(),
            'Nationality':      str(row.get('Nationality') or '').strip(),
            'Location':         str(row.get('Location') or '').strip(),
            'ContractType':     str(row.get('ContractType') or '').strip(),
            'GrossCashMonthly': row.get('GrossCashMonthly') or '',
            'StartDate':        str(row.get('StartDate') or '').strip(),
            'EndDate':          str(row.get('EndDate') or '').strip(),
            'JoiningDate':      str(row.get('JoiningDate') or '').strip(),
            'DateOfBirth':      str(row.get('DateOfBirth') or '').strip(),
            'PrimaryKeyUsed':   'IdentityNumber',
        })
    for r in results:
        nid = r['NormalizedID']
        r['IsDuplicate']    = bool(nid and id_counter[nid] > 1)
        r['DuplicateCount'] = id_counter[nid] if nid else 0
    for r in results:
        if not r['NormalizedID']:
            r['ImportClass'] = 'MissingIdentity'
        elif not r['IDValid']:
            r['ImportClass'] = 'InvalidIdentity'
        elif r['IsDuplicate']:
            r['ImportClass'] = 'NeedsReview'
        else:
            r['ImportClass'] = 'NewEmployee'
    return results, id_counter


def build_report(sheet_name, raw_headers, mapping_detail, results, id_counter, file_path):
    lines = []

    def h(title):
        lines.append(f"\n{'='*72}")
        lines.append(f"  {title}")
        lines.append(f"{'='*72}")

    def row(label, value):
        lines.append(f"  {label:<58} {value}")

    total    = len(results)
    valid_r  = [r for r in results if r['IDValid']]
    dup_r    = [r for r in results if r['IsDuplicate']]
    inv_r    = [r for r in results if r['NormalizedID'] and not r['IDValid']]
    mis_r    = [r for r in results if not r['NormalizedID']]
    saudi_r  = [r for r in results if r['IDType'] == 'Saudi']
    iqama_r  = [r for r in results if r['IDType'] == 'Iqama']
    strange_r= [r for r in results if r['NormalizedID'] and not r['IDValid']
                and r['IDReason'].startswith('starts with')]
    new_r    = [r for r in results if r['ImportClass'] == 'NewEmployee']
    review_r = [r for r in results if r['ImportClass'] == 'NeedsReview']
    dup_ids  = {r['NormalizedID'] for r in dup_r}
    unmapped = [(raw, canon, kind) for raw, canon, kind in mapping_detail if kind == 'unmapped']

    h("FILE INFO")
    row("File:", str(file_path.name))
    row("Sheet name:", sheet_name)
    row("Total columns detected:", str(len(raw_headers)))

    h("1. TOTAL ROWS")
    row("Total data rows:", str(total))

    h("2. SHEET NAME")
    row("Detected sheet:", sheet_name)

    h("3. DETECTED EXCEL COLUMNS (raw)")
    lines.append(f"  {'#':<5} {'Raw Column Header'}")
    lines.append(f"  {'─'*60}")
    for idx, raw in enumerate(raw_headers, 1):
        lines.append(f"  {idx:<5} {raw}")

    h("4. COLUMN MAPPING TABLE  (raw Arabic column → mapped internal field)")
    lines.append(f"  {'Raw Column':<45} {'Internal Field':<30} {'Kind'}")
    lines.append(f"  {'─'*85}")
    for raw, canonical, kind in mapping_detail:
        flag = '  ← UNMAPPED' if kind == 'unmapped' else ''
        lines.append(f"  {raw:<45} {canonical:<30} {kind}{flag}")

    h("5. UNMAPPED COLUMNS")
    row("Unmapped column count:", str(len(unmapped)))
    if unmapped:
        for raw, canon, kind in unmapped:
            lines.append(f"  - {raw}")
    else:
        lines.append("  None — all columns mapped successfully.")

    h("6. IDENTITYNUMBER STATISTICS")
    row("Valid IdentityNumber (10 digits, prefix 1 or 2):", str(len(valid_r)))
    row("  — Saudi National ID (starts with 1):", str(len(saudi_r)))
    row("  — Iqama (starts with 2):", str(len(iqama_r)))
    row("  — Strange prefix (not 1 or 2):", str(len(strange_r)))
    row("Missing IdentityNumber (blank / None):", str(len(mis_r)))
    row("Invalid IdentityNumber (format error):", str(len(inv_r)))
    row("Duplicate IdentityNumber (same ID in >1 row):", str(len(dup_r)))
    row("  — Unique IDs that appear more than once:", str(len(dup_ids)))

    h("7. IMPORT CLASSIFICATION")
    row("NewEmployee (valid, unique IdentityNumber):", str(len(new_r)))
    row("UpdatedEmployee:", "0  (no existing store)")
    row("UnchangedEmployee:", "0  (no existing store)")
    row("NeedsReview (duplicate IdentityNumber in file):", str(len(review_r)))
    row("InvalidIdentity:", str(len(inv_r)))
    row("MissingIdentity:", str(len(mis_r)))

    h("8. MISSING IDENTITYNUMBER — DETAIL")
    if mis_r:
        lines.append(f"  {'Row':<6} {'Name':<45} {'EmpNo':<16} Note")
        lines.append(f"  {'─'*80}")
        for r in mis_r:
            note = "(EmpNo also blank — no key at all)" if not r['EmployeeNumber'] \
                   else "(EmpNo present — secondary only, NOT used for match)"
            lines.append(f"  {r['RowIndex']:<6} {r['Name'][:44]:<45} {r['EmployeeNumber']:<16} {note}")
    else:
        lines.append("  None.")

    h("9. INVALID IDENTITYNUMBER — DETAIL")
    if inv_r:
        lines.append(f"  {'Row':<6} {'RawValue':<25} {'Normalized':<14} {'Reason'}")
        lines.append(f"  {'─'*80}")
        for r in inv_r:
            lines.append(f"  {r['RowIndex']:<6} {str(r['RawIdentityValue'])[:24]:<25} {r['NormalizedID']:<14} {r['IDReason']}")
    else:
        lines.append("  None.")

    h("10. NEEDS REVIEW — TOP 20  (duplicate IdentityNumber within file)")
    if review_r:
        lines.append(f"  {'Row':<6} {'NormalizedID':<14} {'Count':<6} {'Name'}")
        lines.append(f"  {'─'*70}")
        for r in review_r[:20]:
            lines.append(f"  {r['RowIndex']:<6} {r['NormalizedID']:<14} {r['DuplicateCount']:<6} {r['Name']}")
        if len(review_r) > 20:
            lines.append(f"  ... and {len(review_r) - 20} more")
    else:
        lines.append("  None.")

    h("11. PRIMARY KEY CONFIRMATION")
    lines.append("  Every match decision used IdentityNumber as the sole primary key.")
    lines.append("  EmployeeNumber (رمز الموظف) was mapped and stored as secondary reference.")
    lines.append("  EmployeeNumber was NEVER used as a match key.")
    lines.append("  Name (اسم الموظف) was mapped for display only — NOT used for matching.")
    lines.append("")
    emp_only = [r for r in results
                if r['EmployeeNumber'] and (not r['NormalizedID'] or not r['IDValid'])]
    row("Rows with EmpNo but missing/invalid IdentityNumber:", str(len(emp_only)))
    if emp_only:
        lines.append("  These rows go to MissingIdentity or InvalidIdentity — EmpNo NOT used.")
        for r in emp_only:
            lines.append(f"    Row {r['RowIndex']}: EmpNo={r['EmployeeNumber']!r}  Class={r['ImportClass']}")

    h("SUMMARY")
    row("Total rows:", str(total))
    row("Unmapped columns:", str(len(unmapped)))
    row("Valid identities:", str(len(valid_r)))
    row("  Saudi (1xxxxxxxxx):", str(len(saudi_r)))
    row("  Iqama (2xxxxxxxxx):", str(len(iqama_r)))
    row("  Strange prefix:", str(len(strange_r)))
    row("Missing identities:", str(len(mis_r)))
    row("Invalid identities:", str(len(inv_r)))
    row("Duplicate IDs in file:", str(len(dup_r)))
    row("Would-be new employees:", str(len(new_r)))
    row("Needs review:", str(len(review_r)))
    row("Primary key:", "IdentityNumber (الرقم الوطني)")
    row("EmployeeNumber role:", "Secondary — stored, never used as match key")
    row("Name role:", "Display only — NOT used for matching")
    acceptance = "PASS — 0 unmapped columns" if len(unmapped) == 0 \
                 else f"FAIL — {len(unmapped)} unmapped columns remain"
    row("Acceptance condition (0 unmapped):", acceptance)

    return "\n".join(lines)


def safe_print(text):
    try:
        print(text)
    except UnicodeEncodeError:
        print(text.encode('ascii', 'replace').decode('ascii'))


def main():
    if not INPUT_FILE.exists():
        safe_print(f"\nERROR: File not found:\n  {INPUT_FILE}")
        sys.exit(1)

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    safe_print(f"Reading: {INPUT_FILE}")

    sheet_name, raw_headers, mapping_detail, rows = read_excel(INPUT_FILE)
    safe_print(f"Sheet: {sheet_name!r}  |  Rows: {len(rows)}  |  Columns: {len(raw_headers)}")

    results, id_counter = analyze(rows)

    # JSON
    json_out = {
        'sourceFile': str(INPUT_FILE.name),
        'sheetName': sheet_name,
        'totalRows': len(results),
        'columnMapping': [{'raw': r, 'mapped': c, 'kind': k} for r, c, k in mapping_detail],
        'unmappedCount': sum(1 for _, _, k in mapping_detail if k == 'unmapped'),
        'summary': {
            'valid':       sum(1 for r in results if r['IDValid']),
            'saudi':       sum(1 for r in results if r['IDType'] == 'Saudi'),
            'iqama':       sum(1 for r in results if r['IDType'] == 'Iqama'),
            'strange':     sum(1 for r in results if r['NormalizedID'] and not r['IDValid']
                               and r['IDReason'].startswith('starts with')),
            'missing':     sum(1 for r in results if not r['NormalizedID']),
            'invalid':     sum(1 for r in results if r['NormalizedID'] and not r['IDValid']),
            'duplicate':   sum(1 for r in results if r['IsDuplicate']),
            'newEmployee': sum(1 for r in results if r['ImportClass'] == 'NewEmployee'),
            'needsReview': sum(1 for r in results if r['ImportClass'] == 'NeedsReview'),
        },
        'rows': [{
            'rowIndex':         r['RowIndex'],
            'name':             r['Name'],
            'employeeNumber':   r['EmployeeNumber'],
            'rawIdentityValue': str(r['RawIdentityValue']),
            'normalizedID':     r['NormalizedID'],
            'idValid':          r['IDValid'],
            'idType':           r['IDType'],
            'idReason':         r['IDReason'],
            'isDuplicate':      r['IsDuplicate'],
            'duplicateCount':   r['DuplicateCount'],
            'importClass':      r['ImportClass'],
            'primaryKeyUsed':   r['PrimaryKeyUsed'],
            'location':         r['Location'],
            'contractType':     r['ContractType'],
        } for r in results],
    }
    with open(OUT_JSON, 'w', encoding='utf-8') as f:
        json.dump(json_out, f, ensure_ascii=False, indent=2, default=str)
    safe_print(f"JSON:  {OUT_JSON}")

    # Excel
    df_cols = ['RowIndex', 'Name', 'EmployeeNumber', 'RawIdentityValue', 'NormalizedID',
               'IDValid', 'IDType', 'IDReason', 'IsDuplicate', 'DuplicateCount',
               'ImportClass', 'PrimaryKeyUsed', 'Nationality', 'Location',
               'ContractType', 'GrossCashMonthly', 'StartDate', 'EndDate', 'JoiningDate']
    pd.DataFrame([{c: r.get(c, '') for c in df_cols} for r in results])\
      .to_excel(OUT_XLSX, index=False, engine='openpyxl')
    safe_print(f"Excel: {OUT_XLSX}")

    # Report
    report = build_report(sheet_name, raw_headers, mapping_detail,
                          results, id_counter, INPUT_FILE)
    OUT_REPORT.write_text(report, encoding='utf-8')
    safe_print(report)
    safe_print(f"\nReport: {OUT_REPORT}")


if __name__ == '__main__':
    main()
