# -*- coding: utf-8 -*-
"""
link_audit_v2.py
Identity-centric link audit (v2) — Employee Master ↔ Contracts.

Key change from v1:
  EmployeeNumber is NOT a primary key — it's an effective-dated secondary
  identifier that legitimately changes across contract renewals/rehires/cycles.
  IdentityNumber is the only primary person key.

Conceptual entities produced:
  Person                    — unique by IdentityNumber
  EmployeeMasterSnapshot    — one row from the HR Excel
  ContractRecord            — one extracted contract PDF
  EmployeeNumberHistory     — every EmpNo seen for a person, with provenance

Rules:
  - Match key:        IdentityNumber only.
  - EmpNo difference: NOT a conflict — recorded as history entry.
  - Name difference:  display flag only — never used for matching.
  - No DB writes, no UI, no app data modified.

Inputs:
  _contract_lab/employee_master_import_test/inputs/بيانات الموظفين.xlsx
  _contract_lab/outputs/all_contracts_extract_v2_datefix.json

Outputs:
  _contract_lab/outputs/employee_contract_link_audit_v2_identity_model.xlsx
  _contract_lab/outputs/employee_contract_link_audit_v2_identity_model.json
  _contract_lab/outputs/employee_contract_link_audit_v2_identity_model_report.txt
"""

import json, re, sys, datetime as dt
from pathlib import Path
from collections import defaultdict

import openpyxl
import pandas as pd

ROOT             = Path(__file__).resolve().parents[2]
EM_EXCEL         = Path(__file__).parent / "inputs" / "بيانات الموظفين.xlsx"
CONTRACTS_JSON   = ROOT / "_contract_lab" / "outputs" / "all_contracts_extract_v2_datefix.json"
OUT_DIR          = ROOT / "_contract_lab" / "outputs"
OUT_XLSX         = OUT_DIR / "employee_contract_link_audit_v2_identity_model.xlsx"
OUT_JSON         = OUT_DIR / "employee_contract_link_audit_v2_identity_model.json"
OUT_REPORT       = OUT_DIR / "employee_contract_link_audit_v2_identity_model_report.txt"

EM_ALIASES = {
    'الرقم الوطني':       'IdentityNumber',
    'رمز الموظف':         'EmployeeNumber',
    'اسم الموظف':         'Name',
    'إجمالي الراتب':      'GrossCashMonthly',
    'تاريخ بدء العقد':    'StartDate',
    'تاريخ نهاية العقد':  'EndDate',
    'الجنسية':            'Nationality',
    'الموقع':             'Location',
    'نوع العقد':          'ContractType',
    'المسمى الوظيفي':     'Profession',
    'تاريخ التعيين':      'JoiningDate',
    'تاريخ الولادة':      'DateOfBirth',
    'الجنس':              'Gender',
    'العمر':              'Age',
    'مدة الخدمة':         'ServiceDuration',
    'التأمين الصحي':      'HealthInsuranceStatus',
}

SALARY_TOLERANCE = 1.0
PRIORITY_RANK    = {'CRITICAL': 4, 'HIGH': 3, 'MEDIUM': 2, 'LOW': 1, 'NONE': 0}


# ── helpers ───────────────────────────────────────────────────────────────────

def norm_id(value):
    if value is None: return ''
    if isinstance(value, float):
        return '' if value != value else str(int(round(value)))
    if isinstance(value, int):
        return str(value)
    return re.sub(r'[^0-9]', '', str(value).strip())


def validate_id(value):
    id_ = norm_id(value)
    if not id_:        return False, None, 'missing'
    if len(id_) != 10: return False, None, f'length {len(id_)}'
    if id_[0] == '1':  return True, 'Saudi', None
    if id_[0] == '2':  return True, 'Iqama', None
    return False, None, f"starts with '{id_[0]}'"


def norm_num(value):
    if value is None or str(value).strip() in ('', 'None', 'null'): return None
    try:    return float(str(value).replace(',', '').strip())
    except: return None


def norm_date(value):
    if value is None: return ''
    if isinstance(value, (dt.datetime, dt.date)):
        return value.strftime('%Y-%m-%d')
    s = str(value).strip()
    if not s or s in ('None', 'null'): return ''
    m = re.match(r'^(\d{4}-\d{2}-\d{2})\s+\d{2}:\d{2}:\d{2}', s)
    if m: return m.group(1)
    if re.match(r'^\d{4}-\d{2}-\d{2}$', s): return s
    m = re.match(r'^(\d{2})-(\d{2})-(\d{4})$', s)
    if m: return f'{m.group(3)}-{m.group(2)}-{m.group(1)}'
    return s


def norm_emp(value):
    """Normalize EmpNo for comparison (strip leading zeros, lowercase, trim)."""
    return str(value or '').strip().lstrip('0').lower()


def norm_name(value):
    return re.sub(r'\s+', ' ', str(value or '').strip()).lower()


def is_arabic(s):
    return bool(re.search(r'[؀-ۿﭐ-﷿ﹰ-﻿]', str(s or '')))


def date_gap_days(d1, d2):
    if not d1 or not d2: return None
    try:
        return abs((dt.date.fromisoformat(d1[:10]) - dt.date.fromisoformat(d2[:10])).days)
    except Exception:
        return None


# ── readers ───────────────────────────────────────────────────────────────────

def read_employee_master(path):
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows_iter = ws.iter_rows(values_only=True)
    header_row = next(rows_iter, [])
    raw_headers = [str(c or '').strip() for c in header_row]
    mapped = [EM_ALIASES.get(h, h) for h in raw_headers]
    records = []
    for i, data_row in enumerate(rows_iter):
        d = {}
        for col_i, val in enumerate(data_row):
            if col_i < len(mapped):
                d[mapped[col_i]] = val
        records.append({
            'rowIndex':       i + 2,
            'raw_id':         d.get('IdentityNumber'),
            'identityNumber': norm_id(d.get('IdentityNumber')),
            'employeeNumber': str(d.get('EmployeeNumber') or '').strip(),
            'name':           str(d.get('Name') or '').strip(),
            'grossSalary':    norm_num(d.get('GrossCashMonthly')),
            'startDate':      norm_date(d.get('StartDate')),
            'endDate':        norm_date(d.get('EndDate')),
            'joiningDate':    norm_date(d.get('JoiningDate')),
            'dateOfBirth':    norm_date(d.get('DateOfBirth')),
            'nationality':    str(d.get('Nationality') or '').strip(),
            'location':       str(d.get('Location') or '').strip(),
            'profession':     str(d.get('Profession') or '').strip(),
            'contractType':   str(d.get('ContractType') or '').strip(),
            'healthInsuranceStatus': str(d.get('HealthInsuranceStatus') or '').strip(),
            'sourceFile':     path.name,
        })
    wb.close()
    return records


def read_contracts(path):
    with open(path, encoding='utf-8') as f:
        raw = json.load(f)
    out = []
    for c in raw:
        out.append({
            'sourceFile':       c.get('SourceFile', ''),
            'contractNumber':   str(c.get('ContractNumber') or '').strip(),
            'contractVersion':  str(c.get('ContractVersion') or '').strip(),
            'extractionStatus': str(c.get('ExtractionStatus') or '').strip(),
            'raw_id':           c.get('IdentityNumber'),
            'identityNumber':   norm_id(c.get('IdentityNumber')),
            'employeeNumber':   str(c.get('EmployeeNumber') or '').strip(),
            'name':             str(c.get('Name') or '').strip(),
            'basicSalary':      norm_num(c.get('BasicSalary')),
            'grossSalary':      norm_num(c.get('GrossCashMonthly')),
            'startDate':        norm_date(c.get('StartDate')),
            'endDate':          norm_date(c.get('EndDate')),
            'joiningDate':      norm_date(c.get('JoiningDate')),
            'contractEndType':  str(c.get('ContractEndType') or '').strip(),
            'nationality':      str(c.get('Nationality') or '').strip(),
        })
    return out


# ── Person registry ───────────────────────────────────────────────────────────

def build_person_registry(em_records, contract_records):
    """
    Group every record by IdentityNumber.
    Returns:
      persons       — dict id → Person
      unlinked_em   — EM rows with no valid id
      unlinked_con  — contracts with no valid id
    """
    persons      = {}
    unlinked_em  = []
    unlinked_con = []

    def get_or_create(id_, id_type):
        if id_ not in persons:
            persons[id_] = {
                'identityNumber':         id_,
                'idType':                 id_type,
                'emSnapshot':             None,
                'contracts':              [],
                'employeeNumberHistory':  [],   # list of dicts
                'currentName':            '',
                'nationality':            '',
            }
        return persons[id_]

    # EM records
    for r in em_records:
        valid, idtype, _ = validate_id(r['raw_id'])
        if not valid:
            unlinked_em.append(r)
            continue
        p = get_or_create(r['identityNumber'], idtype)
        p['emSnapshot'] = r
        if r['name']:        p['currentName'] = r['name']
        if r['nationality']: p['nationality'] = r['nationality']
        if r['employeeNumber']:
            p['employeeNumberHistory'].append({
                'employeeNumber': r['employeeNumber'],
                'sourceType':     'EmployeeMasterExcel',
                'sourceFile':     r['sourceFile'],
                'contractNumber': '',
                'firstSeenDate':  r['joiningDate'] or r['startDate'] or '',
                'lastSeenDate':   r['endDate'] or '',
                'note':           '',
            })

    # Contract records
    for c in contract_records:
        valid, idtype, _ = validate_id(c['raw_id'])
        if not valid:
            unlinked_con.append(c)
            continue
        p = get_or_create(c['identityNumber'], idtype)
        p['contracts'].append(c)
        if not p['currentName'] and c['name']:
            p['currentName'] = c['name']
        if not p['nationality'] and c['nationality']:
            p['nationality'] = c['nationality']
        if c['employeeNumber']:
            p['employeeNumberHistory'].append({
                'employeeNumber': c['employeeNumber'],
                'sourceType':     'ContractPDF',
                'sourceFile':     c['sourceFile'],
                'contractNumber': c['contractNumber'],
                'firstSeenDate':  c['startDate'] or c['joiningDate'] or '',
                'lastSeenDate':   c['endDate'] or '',
                'note':           '',
            })

    return persons, unlinked_em, unlinked_con


# ── per-person classification ─────────────────────────────────────────────────

def classify_employee_number_history(person):
    """
    Return: list of unique-EmpNo entries with renewal/rehire/cycle hint per record.
    Heuristics for matched persons (EM + ≥1 contract):
      - same EmpNo                    → no history note
      - different EmpNos:
          start gap ≥ 365d            → 'Possible Rehire'
          start gap 30–364d           → 'Possible Renewal'
          start gap < 30d, OPEN_ENDED → 'New Contract Cycle'
          else                        → 'Secondary identifier changed'
    """
    history = person['employeeNumberHistory']
    if not history:
        return []

    # Deduplicate by normalized EmpNo
    seen = {}
    for entry in history:
        key = norm_emp(entry['employeeNumber'])
        if not key:
            continue
        if key not in seen:
            seen[key] = entry.copy()
        else:
            # Merge sources: extend dates, prefer earliest firstSeen / latest lastSeen
            existing = seen[key]
            for fld in ('firstSeenDate', 'lastSeenDate', 'sourceFile', 'contractNumber'):
                if not existing.get(fld) and entry.get(fld):
                    existing[fld] = entry[fld]
    unique = list(seen.values())

    if len(unique) <= 1:
        for entry in unique:
            entry['note'] = ''
            entry['status'] = 'Active'
        return unique

    # Multiple unique EmpNos — annotate based on data
    em      = person['emSnapshot']
    cons    = person['contracts']
    em_sd   = em['startDate'] if em else ''
    con_sd  = cons[0]['startDate'] if cons else ''
    con_cet = cons[0]['contractEndType'] if cons else ''
    gap     = date_gap_days(em_sd, con_sd)

    if gap is not None and gap >= 365:
        note = 'Possible Rehire'
    elif gap is not None and 30 <= gap < 365:
        note = 'Possible Renewal'
    elif gap is not None and gap < 30 and con_cet == 'OPEN_ENDED':
        note = 'New Contract Cycle'
    else:
        note = 'Secondary identifier changed'

    for entry in unique:
        entry['note']   = note
        entry['status'] = 'Historical' if entry['sourceType'] == 'ContractPDF' else 'Active'

    return unique


def classify_person(person):
    """
    Return dict with: presence, issues (list), priority (overall), historyNote.
    Issues are content-only mismatches — EmpNo difference is NOT an issue.
    """
    em       = person['emSnapshot']
    contracts = person['contracts']

    if em and contracts:
        presence = 'BothSides'
    elif em and not contracts:
        presence = 'EmployeeMasterOnly'
    elif contracts and not em:
        presence = 'ContractOnly'
    else:
        presence = 'NoSide'

    issues = []

    # Cross-source consistency only when both sides exist
    if presence == 'BothSides':
        c = contracts[0]   # current data: max 1 contract per person

        # Salary
        em_s, c_s = em['grossSalary'], c['grossSalary']
        if em_s is not None and c_s is not None:
            diff = abs(em_s - c_s)
            if diff > 1000:
                issues.append({'type': 'SalaryMismatch_>1000',  'priority': 'CRITICAL', 'detail': f'diff={diff:.2f}'})
            elif diff > 100:
                issues.append({'type': 'SalaryMismatch_100-1000','priority': 'HIGH',     'detail': f'diff={diff:.2f}'})
            elif diff > SALARY_TOLERANCE:
                pass  # 1–100 SAR — below user-defined threshold; not an issue
        elif em_s is not None and c_s is None and c['extractionStatus'] == 'COMPLETE':
            issues.append({'type': 'SalaryMissingInContract', 'priority': 'HIGH', 'detail': f'EM={em_s}'})

        # StartDate
        gap = date_gap_days(em['startDate'], c['startDate'])
        if gap is not None and gap > 0:
            if gap > 365:
                issues.append({'type': 'StartDateMismatch_>365d', 'priority': 'CRITICAL', 'detail': f'{gap}d'})
            elif gap > 30:
                issues.append({'type': 'StartDateMismatch_>30d',  'priority': 'HIGH',     'detail': f'{gap}d'})
            elif gap >= 2:
                issues.append({'type': 'StartDateMismatch_2-30d', 'priority': 'MEDIUM',   'detail': f'{gap}d'})
            else:
                issues.append({'type': 'StartDateMismatch_1d',    'priority': 'LOW',      'detail': '1d'})

        # EndDate — open-ended aware
        cet = c['contractEndType']
        em_ed, c_ed = em['endDate'], c['endDate']
        if cet == 'OPEN_ENDED' and em_ed:
            issues.append({'type': 'OpenEnded_EM_HasEndDate', 'priority': 'CRITICAL',
                           'detail': f'EM EndDate={em_ed}'})
        elif cet == 'FIXED_TERM':
            if not c_ed and em_ed:
                issues.append({'type': 'FixedTerm_Con_MissingEndDate', 'priority': 'HIGH',
                               'detail': f'EM EndDate={em_ed}'})
            elif c_ed and not em_ed:
                issues.append({'type': 'FixedTerm_EM_MissingEndDate', 'priority': 'HIGH',
                               'detail': f'Con EndDate={c_ed}'})
            else:
                gap = date_gap_days(em_ed, c_ed)
                if gap is not None and gap > 0:
                    if gap > 365:
                        issues.append({'type': 'EndDateMismatch_>365d', 'priority': 'CRITICAL', 'detail': f'{gap}d'})
                    elif gap > 30:
                        issues.append({'type': 'EndDateMismatch_>30d',  'priority': 'HIGH',     'detail': f'{gap}d'})
                    elif gap >= 2:
                        issues.append({'type': 'EndDateMismatch_2-30d', 'priority': 'MEDIUM',   'detail': f'{gap}d'})
                    else:
                        issues.append({'type': 'EndDateMismatch_1d',    'priority': 'LOW',      'detail': '1d'})

        # Name — never used for matching; just a flag
        em_n, c_n = norm_name(em['name']), norm_name(c['name'])
        if em_n and c_n and em_n != c_n:
            if is_arabic(em['name']) != is_arabic(c['name']):
                issues.append({'type': 'NameMismatch_DifferentScript', 'priority': 'MEDIUM'})
            else:
                # Same script — likely Arabic visual-order artifact; LOW
                issues.append({'type': 'NameMismatch_SameScript', 'priority': 'LOW'})

    # Compute overall priority
    overall = 'NONE'
    for iss in issues:
        if PRIORITY_RANK[iss['priority']] > PRIORITY_RANK[overall]:
            overall = iss['priority']

    # Presence-based priority floor
    if presence == 'EmployeeMasterOnly':
        if PRIORITY_RANK['CRITICAL'] > PRIORITY_RANK[overall]:
            overall = 'CRITICAL'
    elif presence == 'ContractOnly':
        if PRIORITY_RANK['CRITICAL'] > PRIORITY_RANK[overall]:
            overall = 'CRITICAL'

    return {
        'presence':           presence,
        'issues':             issues,
        'priority':           overall,
    }


# ── audit driver ──────────────────────────────────────────────────────────────

def run_audit(persons):
    audit_rows = []
    for id_, p in sorted(persons.items()):
        history    = classify_employee_number_history(p)
        cls        = classify_person(p)
        em         = p['emSnapshot']
        cons       = p['contracts']
        c          = cons[0] if cons else None

        unique_emp_count = len(history)
        emp_change_note  = history[0]['note'] if (unique_emp_count > 1 and history) else ''

        audit_rows.append({
            'IdentityNumber':   id_,
            'IDType':           p['idType'],
            'CurrentName':      p['currentName'],
            'Nationality':      p['nationality'],
            'Presence':         cls['presence'],
            'ContractCount':    len(cons),
            'EM_EmpNo':         em['employeeNumber'] if em else '',
            'Con_EmpNo':        c['employeeNumber']  if c  else '',
            'EmpNoChanged':     unique_emp_count > 1,
            'UniqueEmpNoCount': unique_emp_count,
            'EmpHistoryNote':   emp_change_note,
            'EM_GrossSalary':   em['grossSalary']    if em else None,
            'Con_GrossSalary':  c['grossSalary']     if c  else None,
            'EM_StartDate':     em['startDate']      if em else '',
            'Con_StartDate':    c['startDate']       if c  else '',
            'EM_EndDate':       em['endDate']        if em else '',
            'Con_EndDate':      c['endDate']         if c  else '',
            'Con_ContractEndType':  c['contractEndType']  if c else '',
            'Con_ExtractionStatus': c['extractionStatus'] if c else '',
            'Con_Version':      c['contractVersion'] if c  else '',
            'Con_SourceFile':   c['sourceFile']      if c  else '',
            'IssueTypes':       ', '.join(i['type'] for i in cls['issues']),
            'IssueCount':       len(cls['issues']),
            'Priority':         cls['priority'],
            'IssuesDetail':     cls['issues'],
            'EmpNoHistory':     history,
        })
    return audit_rows


# ── report builder ────────────────────────────────────────────────────────────

def build_report(audit_rows, persons, unlinked_em, unlinked_con,
                 em_records, contract_records):
    L = []

    def h(t):
        L.append(f"\n{'='*72}")
        L.append(f"  {t}")
        L.append(f"{'='*72}")

    def r(label, value):
        L.append(f"  {label:<58} {value}")

    total_persons     = len(persons)
    em_only           = [a for a in audit_rows if a['Presence'] == 'EmployeeMasterOnly']
    contract_only     = [a for a in audit_rows if a['Presence'] == 'ContractOnly']
    both_sides        = [a for a in audit_rows if a['Presence'] == 'BothSides']
    multi_emp_persons = [a for a in audit_rows if a['UniqueEmpNoCount'] > 1]
    multi_contracts   = [a for a in audit_rows if a['ContractCount'] > 1]

    # EmpNo categories
    same_id_same_emp = sum(1 for a in both_sides
                           if norm_emp(a['EM_EmpNo']) and norm_emp(a['Con_EmpNo'])
                           and norm_emp(a['EM_EmpNo']) == norm_emp(a['Con_EmpNo']))
    same_id_diff_emp = sum(1 for a in both_sides
                           if norm_emp(a['EM_EmpNo']) and norm_emp(a['Con_EmpNo'])
                           and norm_emp(a['EM_EmpNo']) != norm_emp(a['Con_EmpNo']))
    em_only_emp_no   = sum(1 for a in both_sides
                           if norm_emp(a['EM_EmpNo']) and not norm_emp(a['Con_EmpNo']))
    con_only_emp_no  = sum(1 for a in both_sides
                           if norm_emp(a['Con_EmpNo']) and not norm_emp(a['EM_EmpNo']))
    no_emp_at_all    = sum(1 for a in both_sides
                           if not norm_emp(a['EM_EmpNo']) and not norm_emp(a['Con_EmpNo']))

    # Renewal/rehire breakdown
    renewals  = [a for a in multi_emp_persons if a['EmpHistoryNote'] == 'Possible Renewal']
    rehires   = [a for a in multi_emp_persons if a['EmpHistoryNote'] == 'Possible Rehire']
    cycles    = [a for a in multi_emp_persons if a['EmpHistoryNote'] == 'New Contract Cycle']
    sec_chg   = [a for a in multi_emp_persons if a['EmpHistoryNote'] == 'Secondary identifier changed']

    # Priority counts
    pri_count = {'CRITICAL': 0, 'HIGH': 0, 'MEDIUM': 0, 'LOW': 0, 'NONE': 0}
    for a in audit_rows:
        pri_count[a['Priority']] += 1
    # Add unlinked as CRITICAL (no valid ID at all)
    pri_count['CRITICAL'] += len(unlinked_em) + len(unlinked_con)

    # Issue type counts
    issue_count = defaultdict(int)
    for a in audit_rows:
        for iss in a['IssuesDetail']:
            issue_count[iss['type']] += 1

    # Salary / date mismatch (post-EmpNo-disregard)
    salary_misses = []
    for a in both_sides:
        em_s, c_s = a['EM_GrossSalary'], a['Con_GrossSalary']
        if em_s is not None and c_s is not None:
            diff = abs(em_s - c_s)
            if diff > SALARY_TOLERANCE:
                salary_misses.append({'a': a, 'diff': diff})

    sal_buckets = {'0-100': 0, '101-1000': 0, '1001-5000': 0, '>5000': 0}
    for sm in salary_misses:
        d = sm['diff']
        if d <= 100:    sal_buckets['0-100']     += 1
        elif d <= 1000: sal_buckets['101-1000']  += 1
        elif d <= 5000: sal_buckets['1001-5000'] += 1
        else:           sal_buckets['>5000']     += 1

    # Date mismatch buckets (StartDate + EndDate combined)
    date_buckets = {'1d': 0, '2-30d': 0, '31-365d': 0, '>365d': 0}
    for a in both_sides:
        for sd_pair in [(a['EM_StartDate'], a['Con_StartDate']),
                        (a['EM_EndDate'],   a['Con_EndDate'])]:
            g = date_gap_days(*sd_pair)
            if g is None or g == 0:
                continue
            if g == 1:        date_buckets['1d']      += 1
            elif g <= 30:     date_buckets['2-30d']   += 1
            elif g <= 365:    date_buckets['31-365d'] += 1
            else:             date_buckets['>365d']   += 1

    # Safe auto-link / manual review
    safe_auto    = [a for a in audit_rows if a['Priority'] in ('NONE', 'LOW')
                    and a['Presence'] == 'BothSides']
    manual_review = [a for a in audit_rows if a['Priority'] in ('CRITICAL', 'HIGH', 'MEDIUM')]
    manual_review_count = len(manual_review) + len(unlinked_em) + len(unlinked_con)

    # ── Section 1 ─────────────────────────────────────────────────────────────
    h("1. TOTAL UNIQUE PERSONS BY IDENTITYNUMBER")
    r("Total unique persons (valid IdentityNumber):", str(total_persons))
    r("Saudi (idType=Saudi):", str(sum(1 for p in persons.values() if p['idType']=='Saudi')))
    r("Iqama (idType=Iqama):", str(sum(1 for p in persons.values() if p['idType']=='Iqama')))
    r("EM rows with no valid IdentityNumber (unlinked):", str(len(unlinked_em)))
    r("Contracts with no valid IdentityNumber (unlinked):", str(len(unlinked_con)))

    h("2. PERSONS WITH ONLY EMPLOYEE MASTER")
    r("Count:", str(len(em_only)))

    h("3. PERSONS WITH ONLY CONTRACTS")
    r("Count:", str(len(contract_only)))

    h("4. PERSONS WITH BOTH SIDES")
    r("Count:", str(len(both_sides)))

    h("5. PERSONS WITH MULTIPLE EMPLOYEENUMBERS")
    r("Persons whose UniqueEmpNoCount > 1:", str(len(multi_emp_persons)))
    L.append("")
    L.append("  EmpNo presence breakdown for both-sides persons:")
    r("  Same IdentityNumber + same EmployeeNumber:",  str(same_id_same_emp))
    r("  Same IdentityNumber + different EmployeeNumber:", str(same_id_diff_emp))
    r("  EmpNo only in Excel (Contract has none):", str(em_only_emp_no))
    r("  EmpNo only in Contracts (Excel has none):", str(con_only_emp_no))
    r("  Both sides have no EmpNo:", str(no_emp_at_all))

    h("6. EMPLOYEENUMBER HISTORY CANDIDATES")
    r("Persons with ≥2 unique EmployeeNumbers (history):", str(len(multi_emp_persons)))
    L.append("")
    L.append("  These are NOT failed matches — they are legitimate history entries")
    L.append("  recorded under one Person (keyed by IdentityNumber).")

    h("7. POSSIBLE RENEWAL / RE-HIRE CANDIDATES")
    r("Possible Renewal (start gap 30–364d):", str(len(renewals)))
    r("Possible Rehire (start gap ≥365d):", str(len(rehires)))
    r("New Contract Cycle (gap <30d, OPEN_ENDED contract):", str(len(cycles)))
    r("Secondary identifier changed (other):", str(len(sec_chg)))
    L.append("")
    L.append("  Heuristics applied:")
    L.append("    - start gap ≥ 365d  → Possible Rehire")
    L.append("    - start gap 30–364d → Possible Renewal")
    L.append("    - gap < 30d + OPEN_ENDED → New Contract Cycle")
    L.append("    - else → Secondary identifier changed")

    h("8. CONTRACT COUNT PER PERSON")
    r("Persons with 0 contracts:", str(sum(1 for a in audit_rows if a['ContractCount']==0)))
    r("Persons with 1 contract:",  str(sum(1 for a in audit_rows if a['ContractCount']==1)))
    r("Persons with 2+ contracts:", str(sum(1 for a in audit_rows if a['ContractCount']>=2)))

    h("9. PERSONS WITH MULTIPLE CONTRACTS")
    r("Count:", str(len(multi_contracts)))
    if multi_contracts:
        L.append(f"  {'IdentityNumber':<14} {'ContractCount':<14} {'CurrentName'}")
        L.append(f"  {'─'*60}")
        for a in multi_contracts[:20]:
            L.append(f"  {a['IdentityNumber']:<14} {a['ContractCount']:<14} {a['CurrentName']}")
    else:
        L.append("  None — current dataset has at most 1 contract per IdentityNumber.")

    h("10. SALARY / DATE MISMATCH (EmpNo difference IGNORED as conflict)")
    r("Salary mismatch count (>1 SAR):", str(len(salary_misses)))
    L.append("")
    L.append("  Salary diff distribution:")
    r("    0–100 SAR (informational, below threshold):", str(sal_buckets['0-100']))
    r("    101–1000 SAR (HIGH):",                    str(sal_buckets['101-1000']))
    r("    1001–5000 SAR (CRITICAL):",                str(sal_buckets['1001-5000']))
    r("    >5000 SAR (CRITICAL):",                    str(sal_buckets['>5000']))
    L.append("")
    L.append("  Date diff distribution (StartDate+EndDate combined):")
    r("    1 day (LOW — likely Excel/timezone artifact):", str(date_buckets['1d']))
    r("    2–30 days (MEDIUM):",   str(date_buckets['2-30d']))
    r("    31–365 days (HIGH):",   str(date_buckets['31-365d']))
    r("    >365 days (CRITICAL):", str(date_buckets['>365d']))

    h("PRIORITY DISTRIBUTION")
    r("CRITICAL:", str(pri_count['CRITICAL']))
    r("HIGH:",     str(pri_count['HIGH']))
    r("MEDIUM:",   str(pri_count['MEDIUM']))
    r("LOW:",      str(pri_count['LOW']))
    r("NONE  (no issues):", str(pri_count['NONE']))

    h("ISSUE TYPE COUNTS")
    if issue_count:
        for iss_type, n in sorted(issue_count.items(), key=lambda x: -x[1]):
            r(iss_type, str(n))
    else:
        L.append("  None.")

    h("11. SAFE AUTO-LINK COUNT")
    r("Persons safe for auto-link (BothSides, NONE/LOW priority):", str(len(safe_auto)))
    L.append("  These can be linked without human review:")
    L.append("    - IdentityNumber matched on both sides")
    L.append("    - Salary within 100 SAR")
    L.append("    - Dates within 1 day or matching")
    L.append("    - At most an Arabic visual-order name difference")

    h("12. MANUAL REVIEW COUNT")
    r("Persons requiring manual review:", str(manual_review_count))
    L.append("  Breakdown:")
    r("  CRITICAL persons:", str(pri_count['CRITICAL']))
    r("  HIGH persons:",     str(pri_count['HIGH']))
    r("  MEDIUM persons:",   str(pri_count['MEDIUM']))

    h("TOP 50 CRITICAL CASES")
    crit = [a for a in audit_rows if a['Priority'] == 'CRITICAL']
    L.append(f"  Total CRITICAL person-records: {len(crit)} "
             f"(plus {len(unlinked_em) + len(unlinked_con)} unlinked records below)")
    L.append(f"  {'ID':<14} {'Presence':<22} {'Issues'}")
    L.append(f"  {'─'*100}")
    for a in crit[:50]:
        issues = a['IssueTypes'] if a['IssueTypes'] else '(presence-only)'
        L.append(f"  {a['IdentityNumber']:<14} {a['Presence']:<22} {issues[:60]}")
    if len(crit) > 50:
        L.append(f"  ... and {len(crit)-50} more CRITICAL persons")

    h("13. RECOMMENDED IMPORT DASHBOARD DESIGN")
    L.append("""
  Conceptual data model (4 entities):

    Person (primary key: IdentityNumber)
    ├─ EmployeeMasterSnapshot   (1:1 — replaced on each Excel re-import)
    ├─ ContractRecord           (1:N — accumulates as PDFs are imported)
    └─ EmployeeNumberHistory    (1:N — every EmpNo ever observed)

  ─────────────────────────────────────────────────────────────────────
  Import flows
  ─────────────────────────────────────────────────────────────────────

  A) EMPLOYEE MASTER EXCEL IMPORT
     Input:    بيانات الموظفين.xlsx (full or partial)
     Trigger:  Admin uploads file via "Import Employees" panel
     Stages:
       1. Read file → normalize Arabic columns → validate IdentityNumbers
       2. Preview screen shows:
            New / Updated / Unchanged / Needs Review / Invalid / Missing
            EmpNo history changes (when EmpNo for an existing person differs,
            a history entry is added — NOT flagged as conflict)
       3. Admin reviews + confirms
       4. On commit:
            - Upsert Person by IdentityNumber
            - Replace EmployeeMasterSnapshot
            - Append to EmployeeNumberHistory if EmpNo new for this person
            - Write audit log entries
            - Send invalid/missing IdentityNumber rows to Review Queue

  B) CONTRACT PDF IMPORT
     Input:    1 PDF, several PDFs, a folder, or a ZIP
     Trigger:  User drag-drops or selects files in "Import Contracts" panel
     Stages:
       1. Run extraction pipeline (template classifier → parser)
       2. For each PDF:
            - Validate IdentityNumber
            - Preview shows person match status:
                 a) IdentityNumber matches existing Person  → Add ContractRecord
                 b) IdentityNumber not in registry           → Create ContractOnly Person
                                                               (or Review Queue if confidence low)
                 c) IdentityNumber missing/invalid           → Review Queue
            - If EmpNo on contract differs from EM EmpNo for the same Person,
              add to EmployeeNumberHistory with note (Renewal/Rehire/Cycle)
       3. Admin reviews + confirms
       4. On commit:
            - Append ContractRecord to Person.contracts[]
            - Update EmployeeNumberHistory
            - Write audit log entries

  ─────────────────────────────────────────────────────────────────────
  UI surfaces (high level — no code yet)
  ─────────────────────────────────────────────────────────────────────

  • Import Dashboard
      - Two upload tiles: Employee Master Excel  /  Contract PDFs (single + batch)
      - Preview panel with classification (New, Updated, Unchanged, Review, …)
      - Confirm/cancel buttons
      - Recent imports history (timestamp, source, user, counts)

  • Review Queue
      - Tabs: Missing Identity / Invalid Identity / Ambiguous Match /
              EmpNo History (informational) / Salary Conflict / Date Conflict
      - Per row: side-by-side EM vs Contract, manual resolution actions

  • Employee Profile (per Person)
      - Header: IdentityNumber, current name, nationality, idType
      - Tab "Master snapshot": current EM fields
      - Tab "Contracts timeline": chronological list of ContractRecords
      - Tab "EmployeeNumber history": every EmpNo with source + period + note
      - Tab "Audit log": every change recorded against this person

  ─────────────────────────────────────────────────────────────────────
  Rules enforced everywhere
  ─────────────────────────────────────────────────────────────────────
    - Match key: IdentityNumber only.
    - EmpNo difference: never blocks a match — only a history record.
    - Name difference: display-only flag, never used for matching.
    - Every write produces an audit-log entry.
""".rstrip())

    h("SUMMARY")
    r("Total unique persons (valid IdentityNumber):", str(total_persons))
    r("EM-only persons:",        str(len(em_only)))
    r("Contract-only persons:",  str(len(contract_only)))
    r("Both-sides persons:",     str(len(both_sides)))
    r("Persons with EmpNo history (≥2 unique EmpNos):", str(len(multi_emp_persons)))
    r("  → Possible Renewal:",   str(len(renewals)))
    r("  → Possible Rehire:",    str(len(rehires)))
    r("  → New Contract Cycle:", str(len(cycles)))
    r("  → Secondary identifier changed:", str(len(sec_chg)))
    r("Salary conflicts (>100 SAR):",
      str(sal_buckets['101-1000'] + sal_buckets['1001-5000'] + sal_buckets['>5000']))
    r("Salary conflicts CRITICAL (>1000 SAR):",
      str(sal_buckets['1001-5000'] + sal_buckets['>5000']))
    r("Date conflicts (HIGH+CRITICAL, >30d):",
      str(date_buckets['31-365d'] + date_buckets['>365d']))
    r("Safe auto-link count:",  str(len(safe_auto)))
    r("Manual review count:",   str(manual_review_count))
    r("Unlinked EM (no valid ID):",       str(len(unlinked_em)))
    r("Unlinked contracts (no valid ID):", str(len(unlinked_con)))
    L.append("")
    L.append("  Match key:               IdentityNumber only")
    L.append("  EmployeeNumber:          history record, never a conflict")
    L.append("  Name:                    display flag, never used for matching")

    return "\n".join(L)


# ── main ──────────────────────────────────────────────────────────────────────

def safe_print(s):
    try:
        print(s)
    except UnicodeEncodeError:
        print(s.encode('ascii', 'replace').decode('ascii'))


def main():
    for p, lbl in [(EM_EXCEL, 'EM Excel'), (CONTRACTS_JSON, 'Contracts JSON')]:
        if not Path(p).exists():
            safe_print(f"ERROR: {lbl} not found: {p}")
            sys.exit(1)

    OUT_DIR.mkdir(parents=True, exist_ok=True)

    safe_print(f"Reading EM:        {EM_EXCEL}")
    em_records = read_employee_master(EM_EXCEL)
    safe_print(f"  → {len(em_records)} rows")

    safe_print(f"Reading Contracts: {CONTRACTS_JSON}")
    contract_records = read_contracts(CONTRACTS_JSON)
    safe_print(f"  → {len(contract_records)} records")

    safe_print("Building person registry (keyed by IdentityNumber)...")
    persons, unlinked_em, unlinked_con = build_person_registry(em_records, contract_records)
    safe_print(f"  → {len(persons)} unique persons   "
               f"unlinked EM: {len(unlinked_em)}   unlinked contracts: {len(unlinked_con)}")

    safe_print("Running classification...")
    audit_rows = run_audit(persons)

    # ── JSON ──
    json_out = {
        'inputs': {
            'employeeMaster':  EM_EXCEL.name,
            'contractsFile':   CONTRACTS_JSON.name,
        },
        'totalPersons':         len(persons),
        'totalEMRows':          len(em_records),
        'totalContractRecords': len(contract_records),
        'unlinkedEMRows':       [{'rowIndex': r['rowIndex'],
                                  'name': r['name'],
                                  'employeeNumber': r['employeeNumber'],
                                  'rawId': str(r['raw_id'])}
                                 for r in unlinked_em],
        'unlinkedContracts':    [{'sourceFile': c['sourceFile'],
                                  'name': c['name'],
                                  'employeeNumber': c['employeeNumber'],
                                  'rawId': str(c['raw_id'])}
                                 for c in unlinked_con],
        'persons': [],
    }

    for a in audit_rows:
        json_out['persons'].append({
            'identityNumber':       a['IdentityNumber'],
            'idType':               a['IDType'],
            'currentName':          a['CurrentName'],
            'nationality':          a['Nationality'],
            'presence':             a['Presence'],
            'contractCount':        a['ContractCount'],
            'employeeMasterEmpNo':  a['EM_EmpNo'],
            'contractEmpNo':        a['Con_EmpNo'],
            'uniqueEmpNoCount':     a['UniqueEmpNoCount'],
            'empHistoryNote':       a['EmpHistoryNote'],
            'employeeNumberHistory': a['EmpNoHistory'],
            'em': {
                'grossSalary': a['EM_GrossSalary'],
                'startDate':   a['EM_StartDate'],
                'endDate':     a['EM_EndDate'],
            },
            'contract': {
                'grossSalary':    a['Con_GrossSalary'],
                'startDate':      a['Con_StartDate'],
                'endDate':        a['Con_EndDate'],
                'contractEndType':  a['Con_ContractEndType'],
                'extractionStatus': a['Con_ExtractionStatus'],
                'version':          a['Con_Version'],
                'sourceFile':       a['Con_SourceFile'],
            },
            'priority':    a['Priority'],
            'issueCount':  a['IssueCount'],
            'issues':      a['IssuesDetail'],
        })

    with open(OUT_JSON, 'w', encoding='utf-8') as f:
        json.dump(json_out, f, ensure_ascii=False, indent=2, default=str)
    safe_print(f"JSON:  {OUT_JSON}")

    # ── Excel ──
    flat = []
    for a in audit_rows:
        flat.append({k: v for k, v in a.items() if k not in ('IssuesDetail', 'EmpNoHistory')})

    history_rows = []
    for a in audit_rows:
        for h_entry in a['EmpNoHistory']:
            history_rows.append({
                'IdentityNumber': a['IdentityNumber'],
                'CurrentName':    a['CurrentName'],
                **h_entry,
            })

    with pd.ExcelWriter(OUT_XLSX, engine='openpyxl') as writer:
        pd.DataFrame(flat).to_excel(writer, sheet_name='Persons', index=False)
        pd.DataFrame([f for f in flat if f['Presence']=='EmployeeMasterOnly'])\
          .to_excel(writer, sheet_name='EM_Only', index=False)
        pd.DataFrame([f for f in flat if f['Presence']=='ContractOnly'])\
          .to_excel(writer, sheet_name='Contract_Only', index=False)
        pd.DataFrame([f for f in flat if f['UniqueEmpNoCount']>1])\
          .to_excel(writer, sheet_name='EmpNo_History_Persons', index=False)
        pd.DataFrame(history_rows)\
          .to_excel(writer, sheet_name='EmpNo_History_Detail', index=False)
        pd.DataFrame([f for f in flat if f['Priority']=='CRITICAL'])\
          .to_excel(writer, sheet_name='Priority_Critical', index=False)
        pd.DataFrame([f for f in flat if f['Priority']=='HIGH'])\
          .to_excel(writer, sheet_name='Priority_High', index=False)
        pd.DataFrame([f for f in flat if f['Priority']=='MEDIUM'])\
          .to_excel(writer, sheet_name='Priority_Medium', index=False)
        pd.DataFrame([f for f in flat if f['Priority'] in ('NONE','LOW') and f['Presence']=='BothSides'])\
          .to_excel(writer, sheet_name='Safe_Auto_Link', index=False)
        unlinked_combined = (
            [{'side':'EM','rowIndex':r['rowIndex'],'name':r['name'],
              'employeeNumber':r['employeeNumber'],'rawId':str(r['raw_id'])}
             for r in unlinked_em] +
            [{'side':'Contract','rowIndex':None,'name':c['name'],
              'employeeNumber':c['employeeNumber'],'rawId':str(c['raw_id']),
              'sourceFile':c['sourceFile']}
             for c in unlinked_con]
        )
        if unlinked_combined:
            pd.DataFrame(unlinked_combined).to_excel(writer, sheet_name='Unlinked_NoValidID', index=False)
    safe_print(f"Excel: {OUT_XLSX}")

    # ── Report ──
    report = build_report(audit_rows, persons, unlinked_em, unlinked_con,
                          em_records, contract_records)
    OUT_REPORT.write_text(report, encoding='utf-8')
    safe_print(report)
    safe_print(f"\nReport: {OUT_REPORT}")


if __name__ == '__main__':
    main()
