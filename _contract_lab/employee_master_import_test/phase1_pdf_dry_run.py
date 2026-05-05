# -*- coding: utf-8 -*-
"""
phase1_pdf_dry_run.py
Phase 1 Contract PDF import dry-run — identity-centric.

Mirrors the JS pipeline:
  src/services/imports/contractPdfImportService.js
  src/services/imports/importCommitService.commitContractImport()

Scenario simulated:
  Step A — EM Excel has already been imported (Phase 1 EM dry-run committed).
           The "existing stores" therefore contain 499 persons + 499 history
           entries (one per EM person with non-empty EmpNo).
  Step B — User now imports the contract PDFs (the 437 records that the
           Python extractor produced and that the v2 audit reference uses).

The dry-run compares against the v2 audit numbers in
  _contract_lab/outputs/employee_contract_link_audit_v2_identity_model.json

Inputs:
  EM Excel:    _contract_lab/employee_master_import_test/inputs/بيانات الموظفين.xlsx
  Contracts:   _contract_lab/outputs/all_contracts_extract_v2_datefix.json

Outputs:
  _contract_lab/outputs/phase1_pdf_import_preview.xlsx
  _contract_lab/outputs/phase1_pdf_import_preview.json
  _contract_lab/outputs/phase1_pdf_would_commit.xlsx
  _contract_lab/outputs/phase1_pdf_dry_run_report.txt
"""

import json, re, sys, datetime as dt, uuid
from pathlib import Path
from collections import Counter

import openpyxl
import pandas as pd

ROOT       = Path(__file__).resolve().parents[2]
EM_EXCEL   = Path(__file__).parent / "inputs" / "بيانات الموظفين.xlsx"
CONTRACTS  = ROOT / "_contract_lab" / "outputs" / "all_contracts_extract_v2_datefix.json"
V2_AUDIT   = ROOT / "_contract_lab" / "outputs" / "employee_contract_link_audit_v2_identity_model.json"

OUT_DIR    = ROOT / "_contract_lab" / "outputs"
OUT_PREV_X = OUT_DIR / "phase1_pdf_import_preview.xlsx"
OUT_PREV_J = OUT_DIR / "phase1_pdf_import_preview.json"
OUT_COMMIT = OUT_DIR / "phase1_pdf_would_commit.xlsx"
OUT_REPORT = OUT_DIR / "phase1_pdf_dry_run_report.txt"

# Subset of Arabic aliases needed to read the EM file
ARABIC_ALIASES = {
    'الرقم الوطني':       'IdentityNumber',
    'رمز الموظف':         'EmployeeNumber',
    'اسم الموظف':         'Name',
    'الجنسية':            'Nationality',
    'تاريخ بدء العقد':    'StartDate',
    'تاريخ التعيين':      'JoiningDate',
    'تاريخ نهاية العقد':  'EndDate',
}


# ── helpers (mirror identityModel.js + cleaning.js) ──────────────────────────

def norm_id(value):
    if value is None: return ''
    if isinstance(value, float):
        return '' if value != value else str(int(round(value)))
    if isinstance(value, int):
        return str(value)
    return re.sub(r'[^0-9]', '', str(value).strip())


def validate_id(value):
    id_ = norm_id(value)
    if not id_:        return {'valid': False, 'type': None, 'reason': 'missing'}
    if len(id_) != 10: return {'valid': False, 'type': None, 'reason': f'length {len(id_)}'}
    if id_[0] == '1':  return {'valid': True,  'type': 'Saudi', 'reason': None}
    if id_[0] == '2':  return {'valid': True,  'type': 'Iqama', 'reason': None}
    return {'valid': False, 'type': None, 'reason': f"starts with '{id_[0]}'"}


def norm_emp(value):
    return re.sub(r'\s+', '', str(value or '').strip()).lstrip('0').lower()


def norm_date(value):
    if value is None: return ''
    if isinstance(value, (dt.datetime, dt.date)):
        return value.strftime('%Y-%m-%d')
    s = str(value).strip()
    if not s or s in ('None', 'null'): return ''
    m = re.match(r'^(\d{4}-\d{2}-\d{2})\s+\d{2}:\d{2}:\d{2}', s)
    if m: return m.group(1)
    if re.match(r'^\d{4}-\d{2}-\d{2}$', s): return s
    m = re.match(r'^(\d{2})-(\d{2})-(\d{4})$', s)
    if m: return f'{m.group(3)}-{m.group(2)}-{m.group(1)}'
    return s


def date_gap_days(d1, d2):
    if not d1 or not d2: return None
    try:
        a = dt.date.fromisoformat(d1[:10])
        b = dt.date.fromisoformat(d2[:10])
        return abs((a - b).days)
    except Exception:
        return None


def classify_history_note(em_start, con_start, con_end_type):
    gap = date_gap_days(em_start, con_start)
    if gap is None:           return 'Secondary identifier changed'
    if gap >= 365:            return 'Possible Rehire'
    if gap >= 30:             return 'Possible Renewal'
    if con_end_type == 'OPEN_ENDED': return 'New Contract Cycle'
    return 'Secondary identifier changed'


# ── Step A: Simulate "existing stores" after EM import ───────────────────────

def canonical_em(col):
    raw = str(col or '').strip()
    return ARABIC_ALIASES.get(raw, raw)


def simulate_em_state():
    """Read EM Excel and produce the would-be state of v3 stores AFTER EM import."""
    wb = openpyxl.load_workbook(str(EM_EXCEL), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows_iter = ws.iter_rows(values_only=True)
    header_row = next(rows_iter, [])
    headers = [canonical_em(h) for h in header_row]

    persons = []
    history = []
    snapshots = []
    em_job_id = "phase1-em-job-simulated"

    for i, data_row in enumerate(rows_iter):
        d = {}
        for col_i, val in enumerate(data_row):
            if col_i < len(headers):
                d[headers[col_i]] = val

        raw_id = d.get('IdentityNumber')
        v = validate_id(raw_id)
        if not v['valid']:
            continue

        id_ = norm_id(raw_id)
        emp = str(d.get('EmployeeNumber') or '').strip()
        start_date = norm_date(d.get('StartDate'))
        end_date   = norm_date(d.get('EndDate'))
        join_date  = norm_date(d.get('JoiningDate'))

        persons.append({
            'identityNumber': id_,
            'idType':         v['type'],
            'currentName':    str(d.get('Name') or '').strip(),
            'nationality':    str(d.get('Nationality') or '').strip(),
            'updatedAt':      dt.datetime.now().isoformat(),
        })
        snapshots.append({
            'identityNumber': id_,
            'employeeNumber': emp,
            'startDate':      start_date,
            'endDate':        end_date,
            'joiningDate':    join_date,
        })
        if emp:
            history.append({
                'id':             str(uuid.uuid4()),
                'identityNumber': id_,
                'employeeNumber': emp,
                'sourceType':     'EmployeeMasterExcel',
                'sourceFile':     EM_EXCEL.name,
                'contractNumber': '',
                'firstSeenDate':  join_date or start_date or '',
                'lastSeenDate':   end_date or '',
                'status':         'Active',
                'note':           '',
                'importJobId':    em_job_id,
            })
    wb.close()
    return persons, snapshots, history


# ── Step B: Build contract import preview ────────────────────────────────────

def snapshot_contract_to_record(c, identity, import_job_id, import_date):
    return {
        'id':              str(uuid.uuid4()),
        'identityNumber':  identity,
        'employeeNumber':  str(c.get('EmployeeNumber') or '').strip() or None,
        'contractNumber':  str(c.get('ContractNumber') or '').strip() or None,
        'sourcePdf':       c.get('SourceFile') or None,
        'importJobId':     import_job_id,
        'contractVersion': c.get('ContractVersion') or None,
        'extractionStatus': c.get('ExtractionStatus') or None,
        'startDate':       c.get('StartDate') or None,
        'endDate':         c.get('EndDate') or None,
        'joiningDate':     c.get('JoiningDate') or None,
        'contractEndType': c.get('ContractEndType') or None,
        'basicSalary':     c.get('BasicSalary'),
        'grossCashMonthly': c.get('GrossCashMonthly'),
        'allowances': {
            'HousingAllowance':         c.get('HousingAllowance'),
            'TransportationAllowance':  c.get('TransportationAllowance'),
            'FoodAllowance':            c.get('FoodAllowance'),
            'OTAllowance':              c.get('OTAllowance'),
            'MastersDegreeAllowance':   c.get('MastersDegreeAllowance'),
            'TotalCashAllowances':      c.get('TotalCashAllowances'),
        },
    }


def build_contract_preview(extracted_contracts, existing_persons, existing_history):
    job_id      = f"phase1-pdf-dry-run-{dt.datetime.now().strftime('%Y%m%dT%H%M%S')}"
    import_date = dt.datetime.now().isoformat()

    persons_by_id = {p['identityNumber']: p for p in existing_persons}

    history_by_person = {}
    for h in existing_history:
        history_by_person.setdefault(h['identityNumber'], []).append(h)

    out = {
        'importJobId': job_id,
        'sourceType':  'ContractPDF',
        'generatedAt': import_date,
        'summary': {
            'total': 0,
            'newContractForExistingPerson': 0,
            'newContractOnlyPerson':        0,
            'duplicateContract':            0,
            'empNoHistoryCandidates':       0,
            'needsReview':                  0,
            'invalidIdentity':              0,
            'missingIdentity':              0,
            'extractionError':              0,
        },
        'newContractsForExistingPersons': [],
        'newContractOnlyPersons':         [],
        'duplicateContracts':             [],
        'empNoHistoryCandidates':         [],
        'needsReview':                    [],
        'invalidIdentity':                [],
        'missingIdentity':                [],
        'extractionErrors':               [],
        'auditEntries':                   [],
    }

    # Within-batch duplicate detection
    id_count = Counter()
    for c in extracted_contracts:
        nid = norm_id(c.get('IdentityNumber'))
        if nid: id_count[nid] += 1

    for contract in extracted_contracts:
        out['summary']['total'] += 1

        status = contract.get('ExtractionStatus')
        if status in ('ERROR', 'FAILED_UNKNOWN_TEMPLATE'):
            out['extractionErrors'].append({
                'contract': contract,
                'reason':   f"Extraction error: {contract.get('Error') or 'unknown'}"
                            if status == 'ERROR' else 'Unknown contract template',
            })
            out['summary']['extractionError'] += 1
            continue

        id_ = norm_id(contract.get('IdentityNumber'))
        if not id_:
            out['missingIdentity'].append({'contract': contract, 'reason': 'IdentityNumber blank'})
            out['summary']['missingIdentity'] += 1
            continue

        v = validate_id(id_)
        if not v['valid']:
            out['invalidIdentity'].append({'contract': contract, 'reason': v['reason']})
            out['summary']['invalidIdentity'] += 1
            continue

        if id_count[id_] > 1:
            out['needsReview'].append({
                'contract': contract,
                'reason':   f"IdentityNumber {id_} appears {id_count[id_]}x in this batch",
            })
            out['summary']['needsReview'] += 1
            continue

        new_record = snapshot_contract_to_record({**contract, 'IdentityNumber': id_}, id_, job_id, import_date)
        existing_person = persons_by_id.get(id_)

        if existing_person:
            out['newContractsForExistingPersons'].append({
                'identityNumber': id_,
                'person':         existing_person,
                'contractRecord': new_record,
                'sourceContract': contract,
            })
            out['summary']['newContractForExistingPerson'] += 1
        else:
            out['newContractOnlyPersons'].append({
                'identityNumber': id_,
                'person': {
                    'identityNumber': id_,
                    'idType':         v['type'],
                    'currentName':    str(contract.get('Name') or '').strip(),
                    'nationality':    str(contract.get('Nationality') or '').strip(),
                },
                'contractRecord': new_record,
                'sourceContract': contract,
            })
            out['summary']['newContractOnlyPerson'] += 1

        # EmpNo history candidate.
        # Treat normalize-empty values (e.g. '0000', '   ') as "no EmpNo" —
        # matches v2 audit semantics where norm_emp('0000') == ''.
        emp = str(contract.get('EmployeeNumber') or '').strip()
        emp_norm_val = norm_emp(emp)
        if emp and emp_norm_val:
            existing_hist = history_by_person.get(id_, [])
            known = {norm_emp(e['employeeNumber']) for e in existing_hist if e.get('employeeNumber')}
            known.discard('')   # extra guard against blank-after-normalize entries
            if emp_norm_val not in known:
                note = ''
                if known:
                    prior = sorted(existing_hist, key=lambda h: h.get('firstSeenDate') or '', reverse=True)[0]
                    note = classify_history_note(
                        prior.get('firstSeenDate', ''),
                        contract.get('StartDate', ''),
                        contract.get('ContractEndType', ''),
                    )
                out['empNoHistoryCandidates'].append({
                    'identityNumber':   id_,
                    'newEmpNo':         emp,
                    'previousEmpNumbers': sorted(known),
                    'note':             note,
                    'firstSeen':        len(known) == 0,
                    'entry': {
                        'identityNumber': id_,
                        'employeeNumber': emp,
                        'sourceType':     'ContractPDF',
                        'sourceFile':     contract.get('SourceFile') or '',
                        'contractNumber': str(contract.get('ContractNumber') or '').strip(),
                        'firstSeenDate':  contract.get('StartDate') or contract.get('JoiningDate') or import_date,
                        'lastSeenDate':   contract.get('EndDate') or '',
                        'status':         'Active',
                        'note':           note,
                        'importJobId':    job_id,
                    },
                })
                out['summary']['empNoHistoryCandidates'] += 1
    return out


# ── Would-commit log ─────────────────────────────────────────────────────────

def build_would_commit(preview):
    persons_created = []
    contracts       = []
    history         = []
    audit           = []
    review          = []

    for co in preview['newContractOnlyPersons']:
        persons_created.append(co['person'])

    for item in preview['newContractsForExistingPersons']:
        contracts.append(item['contractRecord'])
        audit.append({
            'action': 'create', 'entityType': 'ContractRecord',
            'entityId': item['contractRecord']['id'],
            'identityNumber': item['identityNumber'],
            'note': 'ContractRecord appended to existing Person',
            'importJobId': preview['importJobId'],
        })
    for item in preview['newContractOnlyPersons']:
        contracts.append(item['contractRecord'])
        audit.append({
            'action': 'create', 'entityType': 'ContractRecord',
            'entityId': item['contractRecord']['id'],
            'identityNumber': item['identityNumber'],
            'note': 'ContractOnly Person + ContractRecord created',
            'importJobId': preview['importJobId'],
        })

    for h in preview['empNoHistoryCandidates']:
        history.append(h['entry'])

    for r in preview['missingIdentity']:
        review.append({
            'reviewType': 'MissingIdentity', 'priority': 'CRITICAL',
            'sourceFile': r['contract'].get('SourceFile'),
            'reason':     r['reason'],
        })
    for r in preview['invalidIdentity']:
        review.append({
            'reviewType': 'InvalidIdentity', 'priority': 'CRITICAL',
            'sourceFile': r['contract'].get('SourceFile'),
            'reason':     r['reason'],
        })
    for r in preview['needsReview']:
        review.append({
            'reviewType': 'AmbiguousMatch', 'priority': 'HIGH',
            'sourceFile': r['contract'].get('SourceFile'),
            'reason':     r['reason'],
        })
    for r in preview['extractionErrors']:
        review.append({
            'reviewType': 'InvalidIdentity', 'priority': 'CRITICAL',
            'sourceFile': r['contract'].get('SourceFile'),
            'reason':     r['reason'],
        })
    for r in preview['duplicateContracts']:
        review.append({
            'reviewType': 'AmbiguousMatch', 'priority': 'MEDIUM',
            'sourceFile': r['contract'].get('SourceFile'),
            'reason':     r['reason'],
        })

    return {
        'persons':   persons_created,
        'contracts': contracts,
        'history':   history,
        'audit':     audit,
        'review':    review,
        'job': {
            'id':         preview['importJobId'],
            'type':       'ContractPDF',
            'status':     'completed',
            'totalItems': preview['summary']['total'],
            'counts':     {**preview['summary'],
                           'personsCreated':   len(persons_created),
                           'contractsWritten': len(contracts),
                           'historyWritten':   len(history),
                           'auditWritten':     len(audit),
                           'reviewWritten':    len(review)},
        },
    }


# ── Acceptance gate ──────────────────────────────────────────────────────────

def run_acceptance_gate(preview, would_commit):
    with open(V2_AUDIT, encoding='utf-8') as f:
        v2 = json.load(f)

    v2_total_contracts   = v2['totalContractRecords']
    v2_unlinked_contracts = len(v2['unlinkedContracts'])
    v2_matched           = sum(1 for p in v2['persons'] if p['presence'] == 'BothSides')
    v2_contract_only     = sum(1 for p in v2['persons'] if p['presence'] == 'ContractOnly')
    v2_empno_history     = sum(1 for p in v2['persons'] if p.get('uniqueEmpNoCount', 1) > 1)

    p1_total          = preview['summary']['total']
    p1_unlinked       = preview['summary']['missingIdentity'] + preview['summary']['invalidIdentity']
    p1_matched        = preview['summary']['newContractForExistingPerson']
    p1_contract_only  = preview['summary']['newContractOnlyPerson']
    p1_history_total  = preview['summary']['empNoHistoryCandidates']
    # v2's "93" specifically counts persons whose normalized EmpNos diverge across
    # sources — i.e. our with-note candidates (existing person had a prior EmpNo
    # AND incoming contract introduces a different one). First-seen entries (no
    # prior EmpNo on file) are NOT what v2 measures.
    p1_history_divergent = sum(
        1 for c in preview['empNoHistoryCandidates'] if c.get('note')
    )
    p1_history_first_seen = sum(
        1 for c in preview['empNoHistoryCandidates'] if not c.get('note')
    )

    rows = [
        ('Total contract records',                v2_total_contracts,    p1_total),
        ('Unlinked contracts (no valid ID)',      v2_unlinked_contracts, p1_unlinked),
        ('Matched contracts (existing Person)',   v2_matched,            p1_matched),
        ('ContractOnly persons',                  v2_contract_only,      p1_contract_only),
        ('EmpNo history candidates — divergence', v2_empno_history,      p1_history_divergent),
    ]
    return rows, {
        'p1_history_total':      p1_history_total,
        'p1_history_first_seen': p1_history_first_seen,
        'p1_history_divergent':  p1_history_divergent,
    }


# ── reporting ────────────────────────────────────────────────────────────────

def build_report(preview, would_commit, acceptance_rows, existing_state, history_breakdown=None):
    L = []

    def h(t):
        L.append(f"\n{'='*72}")
        L.append(f"  {t}")
        L.append(f"{'='*72}")

    def r(label, value):
        L.append(f"  {label:<58} {value}")

    h("PHASE 1 — CONTRACT PDF DRY RUN (after EM import)")
    r("Existing persons (from simulated EM import):", str(len(existing_state['persons'])))
    r("Existing history (from simulated EM import):", str(len(existing_state['history'])))
    r("Source contracts JSON:", CONTRACTS.name)
    r("Note:", "JS parsers deferred to Phase 2; dry-run uses validated Python extraction output")

    s = preview['summary']
    h("PREVIEW SUMMARY")
    r("Total contract records:", str(s['total']))
    r("→ NEW contract for existing Person:", str(s['newContractForExistingPerson']))
    r("→ NEW ContractOnly Person:", str(s['newContractOnlyPerson']))
    r("→ Duplicate contract (same person+number+dates):", str(s['duplicateContract']))
    r("→ EmpNo history candidates:", str(s['empNoHistoryCandidates']))
    r("→ Needs review (within-batch duplicates):", str(s['needsReview']))
    r("→ Invalid identity:", str(s['invalidIdentity']))
    r("→ Missing identity:", str(s['missingIdentity']))
    r("→ Extraction errors:", str(s['extractionError']))

    h("WOULD-COMMIT LOG")
    r("Stores written:", "persons, contractRecords, employeeNumberHistory,")
    L.append(f"  {'':<58} importJobs, importAuditLog, reviewQueue")
    L.append("")
    r("persons (NEW ContractOnly):",         str(len(would_commit['persons'])))
    r("contractRecords (NEW):",              str(len(would_commit['contracts'])))
    r("employeeNumberHistory (NEW):",        str(len(would_commit['history'])))
    r("importAuditLog entries (NEW):",       str(len(would_commit['audit'])))
    r("reviewQueue items (NEW):",            str(len(would_commit['review'])))
    rev_pri = Counter(item['priority'] for item in would_commit['review'])
    for pri, n in sorted(rev_pri.items()):
        r(f"  priority {pri}:", str(n))
    rev_typ = Counter(item['reviewType'] for item in would_commit['review'])
    for typ, n in sorted(rev_typ.items()):
        r(f"  type {typ}:", str(n))
    r("importJobs records (NEW):", "1")

    h("EMPNO HISTORY BREAKDOWN")
    notes = Counter(c.get('note') or '(first seen)' for c in preview['empNoHistoryCandidates'])
    for note, n in sorted(notes.items(), key=lambda x: -x[1]):
        r(f"  {note}:", str(n))
    if history_breakdown:
        L.append("")
        r("Total empNoHistoryCandidates:",          str(history_breakdown['p1_history_total']))
        r("  divergence (with renewal/rehire/cycle/secondary note):",
          str(history_breakdown['p1_history_divergent']))
        r("  first-seen (no prior EmpNo on file):", str(history_breakdown['p1_history_first_seen']))

    h("ACCEPTANCE GATE — Phase 1 vs v2 audit")
    L.append(f"  {'Metric':<42} {'v2 audit':>10} {'phase1':>10} {'match':>8}")
    L.append(f"  {'─'*72}")
    failures = []
    for label, expected, actual in acceptance_rows:
        ok = 'PASS' if expected == actual else 'FAIL'
        if ok == 'FAIL': failures.append(label)
        L.append(f"  {label:<42} {expected:>10} {actual:>10} {ok:>8}")
    L.append("")
    if not failures:
        L.append("  ====== ACCEPTANCE: PASS ======")
    else:
        L.append("  ====== ACCEPTANCE: FAIL ======")
        for f in failures:
            L.append(f"    - {f}")

    return "\n".join(L), failures


# ── main ─────────────────────────────────────────────────────────────────────

def safe_print(s):
    try: print(s)
    except UnicodeEncodeError: print(s.encode('ascii', 'replace').decode('ascii'))


def main():
    for p, lbl in [(EM_EXCEL, 'EM Excel'), (CONTRACTS, 'Contracts JSON'), (V2_AUDIT, 'v2 audit JSON')]:
        if not Path(p).exists():
            safe_print(f"ERROR: {lbl} not found: {p}")
            sys.exit(1)
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    safe_print("Step A: simulating existing-store state from EM Excel...")
    em_persons, em_snapshots, em_history = simulate_em_state()
    safe_print(f"  persons={len(em_persons)}  history={len(em_history)}")

    safe_print(f"Step B: reading {CONTRACTS.name}...")
    with open(CONTRACTS, encoding='utf-8') as f:
        contracts = json.load(f)
    safe_print(f"  contracts={len(contracts)}")

    safe_print("Building contract import preview...")
    preview = build_contract_preview(contracts, em_persons, em_history)
    would_commit = build_would_commit(preview)

    # JSON
    json_payload = {
        'preview': {
            **{k: v for k, v in preview.items() if k != 'auditEntries'},
            'auditEntriesCount': len(preview.get('auditEntries', [])),
        },
        'wouldCommit': {
            'persons':   would_commit['persons'],
            'contractsCount': len(would_commit['contracts']),
            'historyCount':   len(would_commit['history']),
            'auditCount':     len(would_commit['audit']),
            'reviewCount':    len(would_commit['review']),
            'job':            would_commit['job'],
        },
    }
    with open(OUT_PREV_J, 'w', encoding='utf-8') as f:
        json.dump(json_payload, f, ensure_ascii=False, indent=2, default=str)
    safe_print(f"JSON:    {OUT_PREV_J}")

    # Excel — preview rows
    rows_for_xlsx = []
    for item in preview['newContractsForExistingPersons']:
        rows_for_xlsx.append({
            'category': 'NEW_FOR_EXISTING_PERSON',
            'identityNumber': item['identityNumber'],
            'sourcePdf': item['contractRecord']['sourcePdf'],
            'contractNumber': item['contractRecord']['contractNumber'],
            'employeeNumber': item['contractRecord']['employeeNumber'] or '',
        })
    for item in preview['newContractOnlyPersons']:
        rows_for_xlsx.append({
            'category': 'NEW_CONTRACT_ONLY_PERSON',
            'identityNumber': item['identityNumber'],
            'sourcePdf': item['contractRecord']['sourcePdf'],
            'contractNumber': item['contractRecord']['contractNumber'],
            'employeeNumber': item['contractRecord']['employeeNumber'] or '',
        })
    for r in preview['missingIdentity']:
        rows_for_xlsx.append({
            'category': 'MISSING_IDENTITY',
            'sourcePdf': r['contract'].get('SourceFile'),
            'reason': r['reason'],
        })
    for r in preview['invalidIdentity']:
        rows_for_xlsx.append({
            'category': 'INVALID_IDENTITY',
            'sourcePdf': r['contract'].get('SourceFile'),
            'reason': r['reason'],
        })
    pd.DataFrame(rows_for_xlsx).to_excel(OUT_PREV_X, index=False, engine='openpyxl')
    safe_print(f"Excel:   {OUT_PREV_X}")

    # Excel — would-commit per store
    with pd.ExcelWriter(OUT_COMMIT, engine='openpyxl') as writer:
        pd.DataFrame(would_commit['persons']).to_excel(writer, sheet_name='persons_NEW', index=False)
        # Strip nested allowances/rawExtraction for Excel readability
        flat_contracts = []
        for c in would_commit['contracts']:
            flat = {k: v for k, v in c.items() if k not in ('allowances', 'rawExtractionJson')}
            flat['HousingAllowance']        = c.get('allowances', {}).get('HousingAllowance')
            flat['TransportationAllowance'] = c.get('allowances', {}).get('TransportationAllowance')
            flat['TotalCashAllowances']     = c.get('allowances', {}).get('TotalCashAllowances')
            flat_contracts.append(flat)
        pd.DataFrame(flat_contracts).to_excel(writer, sheet_name='contractRecords', index=False)
        pd.DataFrame(would_commit['history']).to_excel(writer, sheet_name='employeeNumberHistory', index=False)
        pd.DataFrame(would_commit['audit']).to_excel(writer, sheet_name='importAuditLog', index=False)
        pd.DataFrame(would_commit['review']).to_excel(writer, sheet_name='reviewQueue', index=False)
        pd.DataFrame([{
            **would_commit['job'],
            'counts': json.dumps(would_commit['job']['counts'], ensure_ascii=False),
        }]).to_excel(writer, sheet_name='importJobs', index=False)
    safe_print(f"Commit:  {OUT_COMMIT}")

    # Acceptance gate + report
    acceptance, history_breakdown = run_acceptance_gate(preview, would_commit)
    report, failures = build_report(
        preview, would_commit, acceptance,
        {'persons': em_persons, 'history': em_history},
        history_breakdown,
    )
    OUT_REPORT.write_text(report, encoding='utf-8')
    safe_print(report)
    safe_print(f"\nReport:  {OUT_REPORT}")

    sys.exit(0 if not failures else 2)


if __name__ == '__main__':
    main()
