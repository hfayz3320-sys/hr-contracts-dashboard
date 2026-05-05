# -*- coding: utf-8 -*-
"""
dry_run_real.py
Dry-run of Employee Master import against the real HR system export file.

Input:   _contract_lab/employee_master_import_test/inputs/بيانات الموظفين.xlsx
Outputs: _contract_lab/outputs/employee_master_real_import_preview.xlsx
         _contract_lab/outputs/employee_master_real_import_preview.json
         _contract_lab/outputs/employee_master_real_import_report.txt

Rules:
  Primary key:    IdentityNumber (الرقم الوطني / رقم الإقامة — 10 digits)
  Secondary key:  EmployeeNumber (only when IdentityNumber missing/invalid)
  Name:           NEVER used for auto-match
  No IndexedDB writes. No app data modified.
"""

import sys, json, re
from pathlib import Path
from collections import Counter

import openpyxl
import pandas as pd

ROOT       = Path(__file__).resolve().parents[2]        # hr-contracts-dashboard-copy/
INPUT_FILE = Path(__file__).parent / "inputs" / "بيانات الموظفين.xlsx"
OUT_DIR    = ROOT / "_contract_lab" / "outputs"
OUT_XLSX   = OUT_DIR / "employee_master_real_import_preview.xlsx"
OUT_JSON   = OUT_DIR / "employee_master_real_import_preview.json"
OUT_REPORT = OUT_DIR / "employee_master_real_import_report.txt"

# ── Arabic column aliases (mirrors schema.js schemaAliasesArabic) ─────────────
ARABIC_ALIASES = {
    'الرقم الوطني':         'IdentityNumber',
    'رقم الهوية الوطنية':   'IdentityNumber',
    'رقم الهوية':           'IdentityNumber',
    'رقم الإقامة':          'IdentityNumber',
    'رقم الاقامة':          'IdentityNumber',
    'الاسم':                 'Name',
    'اسم الموظف':            'Name',
    'الاسم الكامل':          'Name',
    'رقم الموظف':            'EmployeeNumber',
    'الرقم الوظيفي':         'EmployeeNumber',
    'الجنسية':               'Nationality',
    'المسمى الوظيفي':        'Profession',
    'المهنة':                'Profession',
    'تاريخ الميلاد':         'DateOfBirth',
    'تاريخ بداية العقد':     'StartDate',
    'تاريخ المباشرة':        'StartDate',
    'تاريخ نهاية العقد':     'EndDate',
    'تاريخ انتهاء العقد':    'EndDate',
    'تاريخ الانضمام':        'JoiningDate',
    'تاريخ انتهاء الهوية':   'IDExpiryDate',
    'الراتب الأساسي':        'BasicSalary',
    'الراتب':                'BasicSalary',
    'بدل السكن':             'HousingAllowance',
    'بدل النقل':             'TransportationAllowance',
    'بدل الغذاء':            'FoodAllowance',
    'إجمالي البدلات':        'TotalCashAllowances',
    'الراتب الإجمالي':       'GrossCashMonthly',
    'نوع الهوية':            'IDType',
    'رقم الجوال':            'MobileNumber',
    'رقم الهاتف':            'MobileNumber',
    'البريد الإلكتروني':     'Email',
    'رقم الآيبان':           'IBAN',
    'الآيبان':               'IBAN',
    'اسم البنك':             'BankName',
    'الجنس':                 'Gender',
    'الديانة':               'Religion',
    'الحالة الاجتماعية':     'MaritalStatus',
    'المؤهل العلمي':         'Education',
    'التخصص':               'Speciality',
}

# English aliases (normalized lowercase-alphanumeric → canonical field)
_EN_RAW = {
    'sourcefile': 'SourceFile', 'contractnumber': 'ContractNumber',
    'contractno': 'ContractNumber', 'name': 'Name', 'employeename': 'Name',
    'profession': 'Profession', 'jobtitle': 'Profession',
    'employeenumber': 'EmployeeNumber', 'employeeno': 'EmployeeNumber',
    'nationality': 'Nationality', 'dateofbirth': 'DateOfBirth', 'dob': 'DateOfBirth',
    'identitynumber': 'IdentityNumber', 'idnumber': 'IdentityNumber',
    'idtype': 'IDType', 'idexpirydate': 'IDExpiryDate',
    'gender': 'Gender', 'religion': 'Religion', 'maritalstatus': 'MaritalStatus',
    'education': 'Education', 'speciality': 'Speciality', 'specialty': 'Speciality',
    'iban': 'IBAN', 'bankname': 'BankName', 'email': 'Email',
    'mobilenumber': 'MobileNumber', 'phone': 'MobileNumber',
    'contractdurationyears': 'ContractDurationYears',
    'startdate': 'StartDate', 'enddate': 'EndDate', 'joiningdate': 'JoiningDate',
    'basicsalary': 'BasicSalary', 'housingprovided': 'HousingProvided',
    'transportprovided': 'TransportProvided', 'housingallowance': 'HousingAllowance',
    'transportationallowance': 'TransportationAllowance', 'foodallowance': 'FoodAllowance',
    'otallowance': 'OTAllowance', 'mastersdegreeallowance': 'MastersDegreeAllowance',
    'totalcashallowances': 'TotalCashAllowances', 'grosscashmonthly': 'GrossCashMonthly',
}


def canonical_column(col):
    """Return (mapped_field, kind) for a raw Excel header."""
    raw = str(col or '').strip()
    # 1. Arabic exact lookup
    if raw in ARABIC_ALIASES:
        return ARABIC_ALIASES[raw], 'arabic'
    # 2. English normalised lookup
    norm = re.sub(r'[^a-z0-9]', '', raw.lower())
    if norm in _EN_RAW:
        return _EN_RAW[norm], 'english'
    return raw, 'unmapped'


# ── identity helpers ──────────────────────────────────────────────────────────

def normalize_identity(value):
    """Mirrors normalizeIdentityNumber() from cleaning.js."""
    if value is None:
        return ''
    if isinstance(value, float):
        if value != value:          # NaN
            return ''
        return str(int(round(value)))
    if isinstance(value, int):
        return str(value)
    return re.sub(r'[^0-9]', '', str(value).strip())


def validate_identity(value):
    """Mirrors validateIdentityNumber() from cleaning.js."""
    id_ = normalize_identity(value)
    if not id_:
        return {'valid': False, 'type': None, 'reason': 'missing'}
    if len(id_) != 10:
        return {'valid': False, 'type': None, 'reason': f'length {len(id_)} (expected 10)'}
    if id_[0] == '1':
        return {'valid': True,  'type': 'Saudi', 'reason': None}
    if id_[0] == '2':
        return {'valid': True,  'type': 'Iqama', 'reason': None}
    return {'valid': False, 'type': None,
            'reason': f"starts with '{id_[0]}' (expected 1=Saudi or 2=Iqama)"}


# ── Excel reader ──────────────────────────────────────────────────────────────

def read_excel(path):
    """
    Read first sheet; return (sheet_name, raw_headers, mapping_detail, rows).
    Uses openpyxl for raw cell values (avoids scientific-notation float strings).
    """
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    sheet_name = wb.sheetnames[0]
    ws = wb[sheet_name]

    rows_iter = ws.iter_rows(values_only=True)
    header_row = next(rows_iter, None)
    if header_row is None:
        wb.close()
        return sheet_name, [], [], []

    raw_headers    = [str(c or '').strip() for c in header_row]
    mapping_detail = []          # [(raw, canonical, kind), ...]
    mapped_headers = []

    for raw in raw_headers:
        canonical, kind = canonical_column(raw)
        mapped_headers.append(canonical)
        mapping_detail.append((raw, canonical, kind))

    rows = []
    for data_row in rows_iter:
        d = {}
        for col_i, cell_val in enumerate(data_row):
            if col_i >= len(mapped_headers):
                break
            key = mapped_headers[col_i]
            d[key] = cell_val
        rows.append(d)

    wb.close()
    return sheet_name, raw_headers, mapping_detail, rows


# ── analysis ──────────────────────────────────────────────────────────────────

def analyze(rows):
    id_counter = Counter()
    results    = []

    # Pass 1: normalize + validate
    for i, row in enumerate(rows):
        raw_id   = row.get('IdentityNumber', '')
        norm_id  = normalize_identity(raw_id)
        vcheck   = validate_identity(raw_id)

        # EmployeeNumber — collected but NEVER used as primary match key
        raw_emp  = row.get('EmployeeNumber', '')
        norm_emp = str(raw_emp or '').strip()

        if norm_id:
            id_counter[norm_id] += 1

        results.append({
            'RowIndex':        i + 2,
            'RawIdentityValue': raw_id,
            'NormalizedID':    norm_id,
            'IDValid':         vcheck['valid'],
            'IDType':          vcheck['type'] or '',
            'IDReason':        vcheck['reason'] or '',
            'EmployeeNumber':  norm_emp,    # secondary — logged, never matched on
            'Name':            str(row.get('Name') or '').strip(),
            'Nationality':     str(row.get('Nationality') or '').strip(),
            'BasicSalary':     row.get('BasicSalary') or '',
            'StartDate':       str(row.get('StartDate') or '').strip(),
            'EndDate':         str(row.get('EndDate') or '').strip(),
            # Capture which key was ACTUALLY used for matching decision
            'PrimaryKeyUsed':  'IdentityNumber',   # always — EmployeeNumber is never primary
        })

    # Pass 2: tag duplicates
    for r in results:
        nid = r['NormalizedID']
        r['IsDuplicate']    = bool(nid and id_counter[nid] > 1)
        r['DuplicateCount'] = id_counter[nid] if nid else 0

    # Pass 3: classify (no existing store → all valid-unique = New)
    for r in results:
        if not r['NormalizedID']:
            r['ImportClass'] = 'MissingIdentity'
        elif not r['IDValid']:
            r['ImportClass'] = 'InvalidIdentity'
        elif r['IsDuplicate']:
            r['ImportClass'] = 'NeedsReview'
        else:
            r['ImportClass'] = 'NewEmployee'

    return results, id_counter


# ── report builder ────────────────────────────────────────────────────────────

def build_report(sheet_name, raw_headers, mapping_detail, results, id_counter, file_path):
    lines = []

    def h(title):
        lines.append(f"\n{'='*72}")
        lines.append(f"  {title}")
        lines.append(f"{'='*72}")

    def row(label, value):
        lines.append(f"  {label:<58} {value}")

    total    = len(results)
    valid_r  = [r for r in results if r['IDValid']]
    dup_r    = [r for r in results if r['IsDuplicate']]
    inv_r    = [r for r in results if r['NormalizedID'] and not r['IDValid']]
    mis_r    = [r for r in results if not r['NormalizedID']]
    saudi_r  = [r for r in results if r['IDType'] == 'Saudi']
    iqama_r  = [r for r in results if r['IDType'] == 'Iqama']
    strange_r= [r for r in results if r['NormalizedID'] and not r['IDValid']
                and r['IDReason'].startswith('starts with')]
    new_r    = [r for r in results if r['ImportClass'] == 'NewEmployee']
    review_r = [r for r in results if r['ImportClass'] == 'NeedsReview']
    dup_ids  = {r['NormalizedID'] for r in dup_r}

    h("FILE INFO")
    row("File:", str(file_path))
    row("Sheet name:", sheet_name)
    row("Total columns detected:", str(len(raw_headers)))

    h("1. TOTAL ROWS")
    row("Total data rows:", str(total))

    h("2. DETECTED EXCEL COLUMNS (raw)")
    lines.append(f"  {'#':<5} {'Raw Column Header'}")
    lines.append(f"  {'─'*60}")
    for idx, raw in enumerate(raw_headers, 1):
        lines.append(f"  {idx:<5} {raw}")

    h("3. COLUMN MAPPING TABLE  (raw → internal field)")
    lines.append(f"  {'Raw Column':<45} {'Internal Field':<30} {'Kind'}")
    lines.append(f"  {'─'*85}")
    for raw, canonical, kind in mapping_detail:
        flag = '  ← UNMAPPED' if kind == 'unmapped' else ''
        lines.append(f"  {raw:<45} {canonical:<30} {kind}{flag}")

    h("4. IDENTITYNUMBER STATISTICS")
    row("Valid IdentityNumber (10 digits, prefix 1 or 2):", str(len(valid_r)))
    row("  — Saudi National ID (starts with 1):", str(len(saudi_r)))
    row("  — Iqama (starts with 2):", str(len(iqama_r)))
    row("  — Strange prefix (not 1 or 2):", str(len(strange_r)))
    row("Missing IdentityNumber (blank / None):", str(len(mis_r)))
    row("Invalid IdentityNumber (format error):", str(len(inv_r)))
    row("Duplicate IdentityNumber (same ID in >1 row):", str(len(dup_r)))
    row("  — Unique IDs that appear more than once:", str(len(dup_ids)))

    h("5. IMPORT CLASSIFICATION  (no existing store → all valid-unique = New)")
    row("NewEmployee:", str(len(new_r)))
    row("UpdatedEmployee:", "0  (no existing store to compare against)")
    row("UnchangedEmployee:", "0  (no existing store to compare against)")
    row("NeedsReview (duplicate IdentityNumber in this file):", str(len(review_r)))
    row("InvalidIdentity:", str(len(inv_r)))
    row("MissingIdentity:", str(len(mis_r)))

    h("6. MISSING IDENTITYNUMBER — DETAIL")
    if mis_r:
        lines.append(f"  {'Row':<6} {'Name':<45} {'EmployeeNumber':<16} Note")
        lines.append(f"  {'─'*80}")
        for r in mis_r:
            note = "(EmployeeNumber also blank)" if not r['EmployeeNumber'] else "(EmployeeNumber present — secondary only)"
            lines.append(f"  {r['RowIndex']:<6} {r['Name'][:44]:<45} {r['EmployeeNumber']:<16} {note}")
    else:
        lines.append("  None.")

    h("7. INVALID IDENTITYNUMBER — DETAIL")
    if inv_r:
        lines.append(f"  {'Row':<6} {'RawValue':<25} {'Normalized':<14} {'Reason'}")
        lines.append(f"  {'─'*80}")
        for r in inv_r:
            lines.append(f"  {r['RowIndex']:<6} {str(r['RawIdentityValue'])[:24]:<25} {r['NormalizedID']:<14} {r['IDReason']}")
    else:
        lines.append("  None.")

    h("8. STRANGE PREFIX — DETAIL  (not starts with 1 or 2)")
    if strange_r:
        lines.append(f"  {'Row':<6} {'NormalizedID':<14} {'Reason'}")
        lines.append(f"  {'─'*60}")
        for r in strange_r:
            lines.append(f"  {r['RowIndex']:<6} {r['NormalizedID']:<14} {r['IDReason']}")
    else:
        lines.append("  None.")

    h("9. DUPLICATE IDENTITYNUMBER — DETAIL")
    if dup_ids:
        for did in sorted(dup_ids):
            affected = [r for r in results if r['NormalizedID'] == did]
            lines.append(f"\n  ID: {did}  ({len(affected)} rows)")
            for r in affected:
                lines.append(f"    Row {r['RowIndex']:>4}  {r['Name'][:40]:<41} EmpNo={r['EmployeeNumber']}")
    else:
        lines.append("  None.")

    h("10. NEEDS REVIEW — TOP 20  (duplicates flagged for manual resolution)")
    if review_r:
        lines.append(f"  {'Row':<6} {'NormalizedID':<14} {'Count':<6} {'Reason'}")
        lines.append(f"  {'─'*70}")
        for r in review_r[:20]:
            reason = f"Duplicate ID appears {r['DuplicateCount']}x in this file"
            lines.append(f"  {r['RowIndex']:<6} {r['NormalizedID']:<14} {r['DuplicateCount']:<6} {reason}")
        if len(review_r) > 20:
            lines.append(f"  ... and {len(review_r) - 20} more (see preview Excel for full list)")
    else:
        lines.append("  None.")

    h("11. PRIMARY KEY CONFIRMATION")
    lines.append("  Matching decision for every row was made using IdentityNumber only.")
    lines.append("  EmployeeNumber was logged but NEVER used as a match key.")
    lines.append("  Name was not read for matching at all.")
    lines.append("")
    # Count rows where EmployeeNumber was present but IdentityNumber was missing/invalid
    emp_only = [r for r in results
                if r['EmployeeNumber'] and (not r['NormalizedID'] or not r['IDValid'])]
    row("Rows with EmployeeNumber but missing/invalid IdentityNumber:", str(len(emp_only)))
    row("  → These go to MissingIdentity or InvalidIdentity queue.", "")
    row("  → EmployeeNumber was NOT used to match them.", "")

    h("SUMMARY")
    row("Total rows:", str(total))
    row("Valid identities:", str(len(valid_r)))
    row("  Saudi (1xxxxxxxxx):", str(len(saudi_r)))
    row("  Iqama (2xxxxxxxxx):", str(len(iqama_r)))
    row("  Strange prefix:", str(len(strange_r)))
    row("Missing identities:", str(len(mis_r)))
    row("Invalid identities:", str(len(inv_r)))
    row("Duplicate IDs in this file:", str(len(dup_r)))
    row("Would-be new employees:", str(len(new_r)))
    row("Needs review (duplicates):", str(len(review_r)))
    row("Primary key used:", "IdentityNumber (الرقم الوطني / رقم الإقامة)")
    row("EmployeeNumber role:", "Secondary only — never primary match key")
    row("Name role:", "NOT used for matching")

    return "\n".join(lines)


# ── main ──────────────────────────────────────────────────────────────────────

def safe_print(text):
    """Print with ASCII fallback for Windows consoles that don't support Arabic."""
    try:
        print(text)
    except UnicodeEncodeError:
        print(text.encode('ascii', 'replace').decode('ascii'))


def main():
    if not INPUT_FILE.exists():
        safe_print(
            f"\nERROR: Real file not found at:\n"
            f"  {INPUT_FILE}\n\n"
            f"Please place the HR export at that exact path, then rerun this script."
        )
        sys.exit(1)

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    safe_print(f"Reading: {INPUT_FILE}")

    sheet_name, raw_headers, mapping_detail, rows = read_excel(INPUT_FILE)
    safe_print(f"Sheet: {sheet_name!r}  |  Rows: {len(rows)}  |  Columns: {len(raw_headers)}")

    results, id_counter = analyze(rows)

    # ── JSON ──
    json_out = {
        'sourceFile': str(INPUT_FILE.name),
        'sheetName': sheet_name,
        'totalRows': len(results),
        'columnMapping': [
            {'raw': r, 'mapped': c, 'kind': k}
            for r, c, k in mapping_detail
        ],
        'summary': {
            'valid':    sum(1 for r in results if r['IDValid']),
            'saudi':    sum(1 for r in results if r['IDType'] == 'Saudi'),
            'iqama':    sum(1 for r in results if r['IDType'] == 'Iqama'),
            'missing':  sum(1 for r in results if not r['NormalizedID']),
            'invalid':  sum(1 for r in results if r['NormalizedID'] and not r['IDValid']),
            'duplicate': sum(1 for r in results if r['IsDuplicate']),
            'newEmployee': sum(1 for r in results if r['ImportClass'] == 'NewEmployee'),
            'needsReview': sum(1 for r in results if r['ImportClass'] == 'NeedsReview'),
        },
        'rows': [
            {
                'rowIndex':         r['RowIndex'],
                'name':             r['Name'],
                'employeeNumber':   r['EmployeeNumber'],
                'rawIdentityValue': str(r['RawIdentityValue']),
                'normalizedID':     r['NormalizedID'],
                'idValid':          r['IDValid'],
                'idType':           r['IDType'],
                'idReason':         r['IDReason'],
                'isDuplicate':      r['IsDuplicate'],
                'duplicateCount':   r['DuplicateCount'],
                'importClass':      r['ImportClass'],
                'primaryKeyUsed':   r['PrimaryKeyUsed'],
            }
            for r in results
        ],
    }
    with open(OUT_JSON, 'w', encoding='utf-8') as f:
        json.dump(json_out, f, ensure_ascii=False, indent=2, default=str)
    safe_print(f"JSON:  {OUT_JSON}")

    # ── Excel ──
    df_rows = []
    for r in results:
        df_rows.append({
            'RowIndex':       r['RowIndex'],
            'Name':           r['Name'],
            'EmployeeNumber': r['EmployeeNumber'],
            'RawIdentityValue': str(r['RawIdentityValue']),
            'NormalizedID':   r['NormalizedID'],
            'IDValid':        r['IDValid'],
            'IDType':         r['IDType'],
            'IDReason':       r['IDReason'],
            'IsDuplicate':    r['IsDuplicate'],
            'DuplicateCount': r['DuplicateCount'],
            'ImportClass':    r['ImportClass'],
            'PrimaryKeyUsed': r['PrimaryKeyUsed'],
            'Nationality':    r['Nationality'],
            'BasicSalary':    r['BasicSalary'],
            'StartDate':      r['StartDate'],
            'EndDate':        r['EndDate'],
        })
    pd.DataFrame(df_rows).to_excel(OUT_XLSX, index=False, engine='openpyxl')
    safe_print(f"Excel: {OUT_XLSX}")

    # ── Report ──
    report = build_report(sheet_name, raw_headers, mapping_detail,
                          results, id_counter, INPUT_FILE)
    OUT_REPORT.write_text(report, encoding='utf-8')
    safe_print(report)
    safe_print(f"\nReport: {OUT_REPORT}")


if __name__ == '__main__':
    main()
