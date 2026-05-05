# -*- coding: utf-8 -*-
"""
link_audit.py
Employee Master ↔ Contracts linkage audit using IdentityNumber only.

Match key:   IdentityNumber (الرقم الوطني / رقم الإقامة) — 10 digits
Secondary:   EmployeeNumber — logged for mismatch check only, NEVER used to match
Name:        NOT used for matching — only for spelling-difference flag

Inputs:
  Employee Master Excel:  _contract_lab/employee_master_import_test/inputs/بيانات الموظفين.xlsx
  Contracts JSON:         _contract_lab/outputs/all_contracts_extract_v2_datefix.json

Outputs:
  _contract_lab/outputs/employee_contract_link_audit.xlsx
  _contract_lab/outputs/employee_contract_link_audit.json
  _contract_lab/outputs/employee_contract_link_audit_report.txt

Rules:
  - Do not auto-fix anything.
  - Do not import anything.
  - Do not match by EmployeeNumber or Name.
  - Produce audit outputs only.
"""

import json, re, sys, datetime as dt
from pathlib import Path
from collections import Counter, defaultdict

import openpyxl
import pandas as pd

ROOT         = Path(__file__).resolve().parents[2]
EM_EXCEL     = Path(__file__).parent / "inputs" / "بيانات الموظفين.xlsx"
CONTRACTS_JSON = ROOT / "_contract_lab" / "outputs" / "all_contracts_extract_v2_datefix.json"
OUT_DIR      = ROOT / "_contract_lab" / "outputs"
OUT_XLSX     = OUT_DIR / "employee_contract_link_audit.xlsx"
OUT_JSON     = OUT_DIR / "employee_contract_link_audit.json"
OUT_REPORT   = OUT_DIR / "employee_contract_link_audit_report.txt"

# Arabic aliases needed to read Employee Master Excel fields
EM_ALIASES = {
    'الرقم الوطني':   'IdentityNumber',
    'رمز الموظف':     'EmployeeNumber',
    'اسم الموظف':     'Name',
    'إجمالي الراتب':  'GrossCashMonthly',
    'تاريخ بدء العقد': 'StartDate',
    'تاريخ نهاية العقد': 'EndDate',
    'الجنسية':        'Nationality',
    'الموقع':         'Location',
    'نوع العقد':      'ContractType',
    'المسمى الوظيفي': 'Profession',
    'تاريخ التعيين':  'JoiningDate',
}

SALARY_TOLERANCE = 1.0   # SAR — differences ≤ this are considered equal (rounding)


# ── helpers ───────────────────────────────────────────────────────────────────

def norm_id(value):
    """Normalize an identity number to a clean 10-digit string."""
    if value is None:
        return ''
    if isinstance(value, float):
        return '' if value != value else str(int(round(value)))
    if isinstance(value, int):
        return str(value)
    cleaned = re.sub(r'[^0-9]', '', str(value).strip())
    return cleaned


def validate_id(value):
    id_ = norm_id(value)
    if not id_:
        return False, None, 'missing'
    if len(id_) != 10:
        return False, None, f'length {len(id_)}'
    if id_[0] == '1':
        return True, 'Saudi', None
    if id_[0] == '2':
        return True, 'Iqama', None
    return False, None, f"starts with '{id_[0]}'"


def norm_num(value):
    """Normalize a numeric value for comparison."""
    if value is None or str(value).strip() in ('', 'None', 'null'):
        return None
    try:
        return float(str(value).replace(',', '').strip())
    except (ValueError, TypeError):
        return None


def norm_date(value):
    """Return ISO date string (YYYY-MM-DD) or '' from any input."""
    if value is None:
        return ''
    # openpyxl returns Python datetime/date objects for Excel date cells
    if isinstance(value, (dt.datetime, dt.date)):
        return value.strftime('%Y-%m-%d')
    s = str(value).strip()
    if not s or s in ('None', 'null'):
        return ''
    # ISO datetime with time component: "2025-08-01 00:00:00"
    m = re.match(r'^(\d{4}-\d{2}-\d{2})\s+\d{2}:\d{2}:\d{2}', s)
    if m:
        return m.group(1)
    # Already ISO date
    if re.match(r'^\d{4}-\d{2}-\d{2}$', s):
        return s
    # DD-MM-YYYY
    m = re.match(r'^(\d{2})-(\d{2})-(\d{4})$', s)
    if m:
        return f'{m.group(3)}-{m.group(2)}-{m.group(1)}'
    return s


def norm_emp_number(value):
    """Normalize employee number for mismatch comparison."""
    return str(value or '').strip().lstrip('0').lower()


def norm_name(value):
    """Normalize name for mismatch comparison (case + whitespace only)."""
    return re.sub(r'\s+', ' ', str(value or '').strip()).lower()


def is_arabic(s):
    return bool(re.search(r'[؀-ۿﭐ-﷿ﹰ-﻿]', str(s or '')))


# ── readers ───────────────────────────────────────────────────────────────────

def read_employee_master(path):
    """
    Read the real HR Excel. Returns list of dicts with canonical field names.
    Uses openpyxl for raw values (no scientific notation).
    """
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows_iter = ws.iter_rows(values_only=True)
    header_row = next(rows_iter, [])
    raw_headers = [str(c or '').strip() for c in header_row]
    mapped = [EM_ALIASES.get(h, h) for h in raw_headers]

    records = []
    for i, data_row in enumerate(rows_iter):
        d = {}
        for col_i, val in enumerate(data_row):
            if col_i < len(mapped):
                d[mapped[col_i]] = val
        records.append({
            'rowIndex':        i + 2,
            'raw_id':          d.get('IdentityNumber', ''),
            'normalizedID':    norm_id(d.get('IdentityNumber', '')),
            'employeeNumber':  str(d.get('EmployeeNumber') or '').strip(),
            'name':            str(d.get('Name') or '').strip(),
            'grossSalary':     norm_num(d.get('GrossCashMonthly')),
            'startDate':       norm_date(d.get('StartDate')),
            'endDate':         norm_date(d.get('EndDate')),
            'nationality':     str(d.get('Nationality') or '').strip(),
            'location':        str(d.get('Location') or '').strip(),
            'contractType':    str(d.get('ContractType') or '').strip(),
        })
    wb.close()
    return records


def read_contracts(path):
    """Load contracts JSON. Returns list of dicts."""
    with open(path, encoding='utf-8') as f:
        raw = json.load(f)
    records = []
    for c in raw:
        records.append({
            'sourceFile':      c.get('SourceFile', ''),
            'normalizedID':    norm_id(c.get('IdentityNumber')),
            'raw_id':          c.get('IdentityNumber'),
            'employeeNumber':  str(c.get('EmployeeNumber') or '').strip(),
            'name':            str(c.get('Name') or '').strip(),
            'grossSalary':     norm_num(c.get('GrossCashMonthly')),
            'startDate':       norm_date(c.get('StartDate')),
            'endDate':         norm_date(c.get('EndDate')),
            'contractEndType': str(c.get('ContractEndType') or '').strip(),
            'extractionStatus': str(c.get('ExtractionStatus') or '').strip(),
            'contractVersion': str(c.get('ContractVersion') or '').strip(),
            'basicSalary':     norm_num(c.get('BasicSalary')),
            'nationality':     str(c.get('Nationality') or '').strip(),
        })
    return records


# ── audit core ────────────────────────────────────────────────────────────────

def run_audit(em_records, contract_records):
    """
    Link EM ↔ Contracts by IdentityNumber only.
    Returns a list of audit_row dicts — one per unique IdentityNumber
    that appears on either side.
    """

    # Index EM by normalized ID
    em_by_id     = defaultdict(list)
    em_no_id     = []
    for r in em_records:
        valid, _, _ = validate_id(r['raw_id'])
        if valid and r['normalizedID']:
            em_by_id[r['normalizedID']].append(r)
        else:
            em_no_id.append(r)

    # Index contracts by normalized ID
    con_by_id    = defaultdict(list)
    con_no_id    = []
    for c in contract_records:
        valid, _, _ = validate_id(c['raw_id'])
        if valid and c['normalizedID']:
            con_by_id[c['normalizedID']].append(c)
        else:
            con_no_id.append(c)

    # All valid IDs that appear on either side
    all_ids = set(em_by_id.keys()) | set(con_by_id.keys())

    audit_rows = []

    for id_ in sorted(all_ids):
        em_list  = em_by_id.get(id_, [])
        con_list = con_by_id.get(id_, [])

        em  = em_list[0]  if em_list  else None
        con = con_list[0] if con_list else None

        row = {
            'IdentityNumber':    id_,
            'IDType':            ('Saudi' if id_[0] == '1' else 'Iqama'),
            'EM_Present':        bool(em),
            'Contract_Present':  bool(con),
            'EM_DuplicateCount': len(em_list),
            'Con_DuplicateCount': len(con_list),
            # EM fields
            'EM_RowIndex':       em['rowIndex']      if em else None,
            'EM_Name':           em['name']           if em else '',
            'EM_EmpNo':          em['employeeNumber'] if em else '',
            'EM_GrossSalary':    em['grossSalary']    if em else None,
            'EM_StartDate':      em['startDate']      if em else '',
            'EM_EndDate':        em['endDate']        if em else '',
            'EM_Location':       em['location']       if em else '',
            'EM_ContractType':   em['contractType']   if em else '',
            # Contract fields
            'Con_SourceFile':    con['sourceFile']       if con else '',
            'Con_Name':          con['name']              if con else '',
            'Con_EmpNo':         con['employeeNumber']    if con else '',
            'Con_GrossSalary':   con['grossSalary']       if con else None,
            'Con_StartDate':     con['startDate']         if con else '',
            'Con_EndDate':       con['endDate']           if con else '',
            'Con_ContractEndType': con['contractEndType'] if con else '',
            'Con_ExtractionStatus': con['extractionStatus'] if con else '',
            'Con_Version':       con['contractVersion']   if con else '',
            'Con_BasicSalary':   con['basicSalary']       if con else None,
        }

        # ── mismatch checks ──────────────────────────────────────────────────

        flags = []

        if not em:
            row['LinkStatus'] = 'ContractNoEmployee'
            flags.append('No matching Employee Master record')
        elif not con:
            row['LinkStatus'] = 'EmployeeNoContract'
            flags.append('No matching Contract record')
        else:
            row['LinkStatus'] = 'Matched'

            # EmployeeNumber mismatch (logged, not used for matching)
            em_emp  = norm_emp_number(em['employeeNumber'])
            con_emp = norm_emp_number(con['employeeNumber'])
            if em_emp and con_emp and em_emp != con_emp:
                flags.append(f"EmpNo mismatch: EM={em['employeeNumber']} vs Con={con['employeeNumber']}")
                row['EmpNoMismatch'] = True
            else:
                row['EmpNoMismatch'] = False

            # Name mismatch
            em_name  = norm_name(em['name'])
            con_name = norm_name(con['name'])
            if em_name and con_name and em_name != con_name:
                script_note = 'different-script' if (is_arabic(em['name']) != is_arabic(con['name'])) \
                              else 'same-script-different-value'
                flags.append(f"Name mismatch ({script_note}): EM={em['name']!r} vs Con={con['name']!r}")
                row['NameMismatch'] = script_note
            else:
                row['NameMismatch'] = ''

            # Salary mismatch
            em_sal  = em['grossSalary']
            con_sal = con['grossSalary']
            if em_sal is not None and con_sal is not None:
                diff = abs(em_sal - con_sal)
                if diff > SALARY_TOLERANCE:
                    flags.append(f"Salary mismatch: EM={em_sal} vs Con={con_sal} (diff={diff:.2f})")
                    row['SalaryMismatch'] = True
                    row['SalaryDiff']     = round(diff, 2)
                else:
                    row['SalaryMismatch'] = False
                    row['SalaryDiff']     = round(diff, 2)
            elif em_sal is None and con_sal is not None:
                row['SalaryMismatch'] = False   # EM has no salary data (e.g. Arabic-only)
                row['SalaryDiff']     = None
            elif em_sal is not None and con_sal is None:
                flags.append(f"Salary present in EM ({em_sal}) but missing in Contract")
                row['SalaryMismatch'] = True
                row['SalaryDiff']     = None
            else:
                row['SalaryMismatch'] = False
                row['SalaryDiff']     = None

            # Date mismatch: StartDate
            em_sd  = em['startDate']
            con_sd = con['startDate']
            start_flag = ''
            if em_sd and con_sd and em_sd != con_sd:
                start_flag = f"StartDate mismatch: EM={em_sd} vs Con={con_sd}"
                flags.append(start_flag)
            row['StartDateMismatch'] = bool(start_flag)

            # Date mismatch: EndDate — open-ended aware
            em_ed   = em['endDate']
            con_ed  = con['endDate']
            con_cet = con['contractEndType']
            end_flag = ''
            if con_cet == 'OPEN_ENDED':
                if not em_ed:
                    end_flag = ''          # both agree: no end date → OK
                else:
                    end_flag = f"Contract is OPEN_ENDED but EM has EndDate={em_ed}"
                    flags.append(end_flag)
            else:
                if em_ed and con_ed and em_ed != con_ed:
                    end_flag = f"EndDate mismatch: EM={em_ed} vs Con={con_ed}"
                    flags.append(end_flag)
                elif em_ed and not con_ed:
                    end_flag = f"EM has EndDate={em_ed} but Contract has none"
                    flags.append(end_flag)
                elif con_ed and not em_ed:
                    end_flag = f"Contract has EndDate={con_ed} but EM has none"
                    flags.append(end_flag)
            row['EndDateMismatch'] = bool(end_flag)

            # Duplicate in either side
            if len(em_list) > 1:
                flags.append(f"Duplicate in EM: {len(em_list)} EM rows share this ID")
            if len(con_list) > 1:
                flags.append(f"Duplicate in Contracts: {len(con_list)} contracts share this ID")

        row['Flags']     = ' | '.join(flags) if flags else ''
        row['FlagCount'] = len(flags)
        row.setdefault('EmpNoMismatch',    False)
        row.setdefault('NameMismatch',     '')
        row.setdefault('SalaryMismatch',   False)
        row.setdefault('SalaryDiff',       None)
        row.setdefault('StartDateMismatch', False)
        row.setdefault('EndDateMismatch',   False)
        audit_rows.append(row)

    # ── records with no valid ID (can't be linked) ───────────────────────────
    unlinked_em  = [{'side': 'EM',       'rowIndex': r['rowIndex'],
                     'name': r['name'],  'employeeNumber': r['employeeNumber'],
                     'raw_id': str(r['raw_id']), 'reason': 'missing/invalid IdentityNumber'}
                    for r in em_no_id]
    unlinked_con = [{'side': 'Contract', 'sourceFile': c['sourceFile'],
                     'name': c['name'],  'employeeNumber': c['employeeNumber'],
                     'raw_id': str(c['raw_id']), 'reason': 'missing/invalid IdentityNumber'}
                    for c in con_no_id]

    return audit_rows, unlinked_em, unlinked_con


# ── report ────────────────────────────────────────────────────────────────────

def build_report(audit_rows, unlinked_em, unlinked_con,
                 em_records, contract_records):
    lines = []

    def h(title):
        lines.append(f"\n{'='*72}")
        lines.append(f"  {title}")
        lines.append(f"{'='*72}")

    def row(label, value):
        lines.append(f"  {label:<58} {value}")

    # ── basic counts ──────────────────────────────────────────────────────────
    em_total   = len(em_records)
    em_valid   = sum(1 for r in em_records if validate_id(r['raw_id'])[0])
    em_missing = sum(1 for r in em_records if not norm_id(r['raw_id']))
    em_invalid = sum(1 for r in em_records
                     if norm_id(r['raw_id']) and not validate_id(r['raw_id'])[0])

    con_total  = len(contract_records)
    con_valid  = sum(1 for c in contract_records if validate_id(c['raw_id'])[0])

    matched        = [r for r in audit_rows if r['LinkStatus'] == 'Matched']
    em_no_contract = [r for r in audit_rows if r['LinkStatus'] == 'EmployeeNoContract']
    con_no_em      = [r for r in audit_rows if r['LinkStatus'] == 'ContractNoEmployee']

    em_dup_ids  = {r['IdentityNumber'] for r in audit_rows if r['EM_DuplicateCount']  > 1}
    con_dup_ids = {r['IdentityNumber'] for r in audit_rows if r['Con_DuplicateCount'] > 1}

    empno_mismatch   = [r for r in matched if r['EmpNoMismatch']]
    name_mismatch    = [r for r in matched if r['NameMismatch']]
    salary_mismatch  = [r for r in matched if r['SalaryMismatch']]
    start_mismatch   = [r for r in matched if r['StartDateMismatch']]
    end_mismatch     = [r for r in matched if r['EndDateMismatch']]

    h("1. TOTAL EMPLOYEE MASTER ROWS")
    row("Total EM rows:", str(em_total))

    h("2. VALID EMPLOYEE MASTER IDENTITIES")
    saudi_em = sum(1 for r in em_records if validate_id(r['raw_id'])[1] == 'Saudi')
    iqama_em = sum(1 for r in em_records if validate_id(r['raw_id'])[1] == 'Iqama')
    row("Valid IdentityNumber (10 digits, prefix 1 or 2):", str(em_valid))
    row("  — Saudi National ID (starts with 1):", str(saudi_em))
    row("  — Iqama (starts with 2):", str(iqama_em))

    h("3. MISSING / INVALID EMPLOYEE MASTER IDENTITIES")
    row("Missing IdentityNumber (blank):", str(em_missing))
    row("Invalid IdentityNumber (format error):", str(em_invalid))
    row("  (these rows CANNOT be linked — unlinked EM list below)", "")

    h("4. TOTAL CONTRACT RECORDS")
    row("Total contracts:", str(con_total))

    h("5. VALID CONTRACT IDENTITIES")
    saudi_con = sum(1 for c in contract_records if validate_id(c['raw_id'])[1] == 'Saudi')
    iqama_con = sum(1 for c in contract_records if validate_id(c['raw_id'])[1] == 'Iqama')
    row("Valid IdentityNumber:", str(con_valid))
    row("  — Saudi National ID:", str(saudi_con))
    row("  — Iqama:", str(iqama_con))
    row("  — Missing/invalid in Contracts:", str(con_total - con_valid))

    h("6. EMPLOYEE MASTER MATCHED TO CONTRACT")
    row("Matched (EM + Contract share same IdentityNumber):", str(len(matched)))

    h("7. EMPLOYEE MASTER WITHOUT CONTRACT")
    row("EM rows with valid ID but no matching Contract:", str(len(em_no_contract)))

    h("8. CONTRACT WITHOUT EMPLOYEE MASTER")
    row("Contracts with valid ID but no matching EM row:", str(len(con_no_em)))

    h("9. DUPLICATE IDENTITYNUMBER — EMPLOYEE MASTER")
    row("Unique IDs that appear more than once in EM:", str(len(em_dup_ids)))
    if em_dup_ids:
        for did in sorted(em_dup_ids):
            group = [r for r in audit_rows if r['IdentityNumber'] == did]
            if group:
                lines.append(f"  {did}  (EM count: {group[0]['EM_DuplicateCount']})")
    else:
        lines.append("  None.")

    h("10. DUPLICATE IDENTITYNUMBER — CONTRACTS")
    row("Unique IDs that appear more than once in Contracts:", str(len(con_dup_ids)))
    if con_dup_ids:
        for did in sorted(con_dup_ids):
            group = [r for r in audit_rows if r['IdentityNumber'] == did]
            if group:
                lines.append(f"  {did}  (Contract count: {group[0]['Con_DuplicateCount']})")
    else:
        lines.append("  None.")

    h("11. EMPLOYEENUMBER MISMATCH")
    row("Both sides have EmpNo but values differ:", str(len(empno_mismatch)))
    if empno_mismatch:
        lines.append(f"  {'ID':<14} {'EM EmpNo':<12} {'Contract EmpNo'}")
        lines.append(f"  {'─'*50}")
        for r in empno_mismatch[:20]:
            lines.append(f"  {r['IdentityNumber']:<14} {r['EM_EmpNo']:<12} {r['Con_EmpNo']}")
        if len(empno_mismatch) > 20:
            lines.append(f"  ... and {len(empno_mismatch)-20} more")

    h("12. NAME MISMATCH / POSSIBLE SPELLING DIFFERENCE")
    row("Name differs between EM and Contract:", str(len(name_mismatch)))
    if name_mismatch:
        lines.append(f"  {'ID':<14} {'Script':<22} {'EM Name':<35} {'Contract Name'}")
        lines.append(f"  {'─'*95}")
        for r in name_mismatch[:20]:
            lines.append(f"  {r['IdentityNumber']:<14} {r['NameMismatch']:<22} "
                         f"{r['EM_Name'][:34]:<35} {r['Con_Name'][:34]}")
        if len(name_mismatch) > 20:
            lines.append(f"  ... and {len(name_mismatch)-20} more")

    h("13. SALARY MISMATCH  (EM GrossCashMonthly vs Contract GrossCashMonthly)")
    row("Tolerance applied:", f"{SALARY_TOLERANCE} SAR")
    row("Salary mismatch count:", str(len(salary_mismatch)))
    if salary_mismatch:
        lines.append(f"  {'ID':<14} {'EM Salary':>12} {'Con Salary':>12} {'Diff':>10}")
        lines.append(f"  {'─'*55}")
        for r in sorted(salary_mismatch, key=lambda x: -(x['SalaryDiff'] or 0))[:20]:
            lines.append(f"  {r['IdentityNumber']:<14} "
                         f"{str(r['EM_GrossSalary']):>12} "
                         f"{str(r['Con_GrossSalary']):>12} "
                         f"{str(r['SalaryDiff']):>10}")
        if len(salary_mismatch) > 20:
            lines.append(f"  ... and {len(salary_mismatch)-20} more")

    h("14. CONTRACT DATE MISMATCH  (StartDate / EndDate)")
    row("StartDate mismatch count:", str(len(start_mismatch)))
    if start_mismatch:
        lines.append(f"  {'ID':<14} {'EM Start':<14} {'Con Start'}")
        lines.append(f"  {'─'*45}")
        for r in start_mismatch[:10]:
            lines.append(f"  {r['IdentityNumber']:<14} {r['EM_StartDate']:<14} {r['Con_StartDate']}")
    row("EndDate mismatch count:", str(len(end_mismatch)))
    if end_mismatch:
        lines.append(f"  {'ID':<14} {'EM End':<14} {'Con End':<14} {'ContractEndType'}")
        lines.append(f"  {'─'*60}")
        for r in end_mismatch[:10]:
            lines.append(f"  {r['IdentityNumber']:<14} {r['EM_EndDate']:<14} "
                         f"{r['Con_EndDate']:<14} {r['Con_ContractEndType']}")

    h("15. OPEN-ENDED CONTRACT HANDLING")
    open_ended = [r for r in matched if r['Con_ContractEndType'] == 'OPEN_ENDED']
    oe_ok      = [r for r in open_ended if not r['EM_EndDate']]
    oe_review  = [r for r in open_ended if r['EM_EndDate']]
    row("Open-ended contracts (ContractEndType=OPEN_ENDED):", str(len(open_ended)))
    row("  OK — both sides have no EndDate:", str(len(oe_ok)))
    row("  Review — Contract OPEN_ENDED but EM has EndDate:", str(len(oe_review)))
    if oe_review:
        lines.append(f"  {'ID':<14} {'EM EndDate'}")
        lines.append(f"  {'─'*30}")
        for r in oe_review:
            lines.append(f"  {r['IdentityNumber']:<14} {r['EM_EndDate']}")

    h("16. TOP 50 REVIEW CASES WITH REASON")
    # Collect all review-worthy rows in priority order
    review_pool = []
    # 1. EM missing/invalid ID (can't link)
    for r in unlinked_em:
        review_pool.append({'ID': r['raw_id'], 'Side': 'EM',
                            'Category': 'MissingOrInvalidID',
                            'Reason': f"EM row {r['rowIndex']} — {r['reason']}",
                            'Name': r['name'], 'EmpNo': r['employeeNumber']})
    # 2. Contract missing/invalid ID
    for c in unlinked_con:
        review_pool.append({'ID': c['raw_id'], 'Side': 'Contract',
                            'Category': 'MissingOrInvalidID',
                            'Reason': f"Contract {c['sourceFile']} — {c['reason']}",
                            'Name': c['name'], 'EmpNo': c['employeeNumber']})
    # 3. No matching contract
    for r in em_no_contract:
        review_pool.append({'ID': r['IdentityNumber'], 'Side': 'EM',
                            'Category': 'NoMatchingContract',
                            'Reason': 'EM has no matching Contract',
                            'Name': r['EM_Name'], 'EmpNo': r['EM_EmpNo']})
    # 4. No matching employee
    for r in con_no_em:
        review_pool.append({'ID': r['IdentityNumber'], 'Side': 'Contract',
                            'Category': 'NoMatchingEmployee',
                            'Reason': f"Contract ({r['Con_SourceFile']}) has no matching EM",
                            'Name': r['Con_Name'], 'EmpNo': r['Con_EmpNo']})
    # 5. Matched but with flags (salary, date, empno, name)
    flagged = sorted([r for r in matched if r['FlagCount'] > 0],
                     key=lambda x: -x['FlagCount'])
    for r in flagged:
        review_pool.append({'ID': r['IdentityNumber'], 'Side': 'Both',
                            'Category': 'MatchedWithFlags',
                            'Reason': r['Flags'],
                            'Name': r['EM_Name'], 'EmpNo': r['EM_EmpNo']})

    lines.append(f"  Total review cases: {len(review_pool)}")
    lines.append(f"  {'ID':<14} {'Category':<25} {'Side':<10} {'Name':<35} Reason")
    lines.append(f"  {'─'*110}")
    for item in review_pool[:50]:
        lines.append(f"  {str(item['ID'])[:13]:<14} {item['Category']:<25} {item['Side']:<10} "
                     f"{str(item['Name'])[:34]:<35} {item['Reason'][:60]}")
    if len(review_pool) > 50:
        lines.append(f"  ... and {len(review_pool)-50} more review cases (see audit Excel/JSON)")

    h("SUMMARY")
    row("Total EM rows:", str(em_total))
    row("Total Contract records:", str(con_total))
    row("Valid EM identities:", str(em_valid))
    row("Valid Contract identities:", str(con_valid))
    row("Matched (IdentityNumber found on both sides):", str(len(matched)))
    row("EM without matching Contract:", str(len(em_no_contract)))
    row("Contract without matching EM:", str(len(con_no_em)))
    row("EM unlinked (no valid IdentityNumber):", str(len(unlinked_em)))
    row("Contract unlinked (no valid IdentityNumber):", str(len(unlinked_con)))
    row("Duplicate IDs in EM:", str(len(em_dup_ids)))
    row("Duplicate IDs in Contracts:", str(len(con_dup_ids)))
    row("EmployeeNumber mismatch:", str(len(empno_mismatch)))
    row("Name mismatch:", str(len(name_mismatch)))
    row("Salary mismatch (>" + str(SALARY_TOLERANCE) + " SAR):", str(len(salary_mismatch)))
    row("StartDate mismatch:", str(len(start_mismatch)))
    row("EndDate mismatch:", str(len(end_mismatch)))
    row("Open-ended: OK (both sides no EndDate):", str(len(oe_ok)))
    row("Open-ended: Review (EM has EndDate):", str(len(oe_review)))
    row("Total review cases:", str(len(review_pool)))
    row("Match key used:", "IdentityNumber only")
    row("EmployeeNumber used for matching:", "NO — secondary/mismatch-check only")
    row("Name used for matching:", "NO — mismatch-flag only")

    return "\n".join(lines)


# ── main ──────────────────────────────────────────────────────────────────────

def safe_print(text):
    try:
        print(text)
    except UnicodeEncodeError:
        print(text.encode('ascii', 'replace').decode('ascii'))


def main():
    for p, label in [(EM_EXCEL, 'Employee Master Excel'),
                     (CONTRACTS_JSON, 'Contracts JSON')]:
        if not Path(p).exists():
            safe_print(f"ERROR: {label} not found:\n  {p}")
            sys.exit(1)

    OUT_DIR.mkdir(parents=True, exist_ok=True)

    safe_print(f"Reading Employee Master: {EM_EXCEL}")
    em_records = read_employee_master(EM_EXCEL)
    safe_print(f"  → {len(em_records)} rows")

    safe_print(f"Reading Contracts:       {CONTRACTS_JSON}")
    contract_records = read_contracts(CONTRACTS_JSON)
    safe_print(f"  → {len(contract_records)} records")

    safe_print("Running linkage audit...")
    audit_rows, unlinked_em, unlinked_con = run_audit(em_records, contract_records)

    # ── JSON output ──
    json_out = {
        'inputs': {
            'employeeMaster': str(EM_EXCEL.name),
            'contractsFile':  str(CONTRACTS_JSON.name),
        },
        'summary': {
            'emTotal':            len(em_records),
            'contractsTotal':     len(contract_records),
            'matched':            sum(1 for r in audit_rows if r['LinkStatus'] == 'Matched'),
            'emNoContract':       sum(1 for r in audit_rows if r['LinkStatus'] == 'EmployeeNoContract'),
            'contractNoEmployee': sum(1 for r in audit_rows if r['LinkStatus'] == 'ContractNoEmployee'),
            'emUnlinked':         len(unlinked_em),
            'contractUnlinked':   len(unlinked_con),
            'salaryMismatch':     sum(1 for r in audit_rows if r.get('SalaryMismatch')),
            'startDateMismatch':  sum(1 for r in audit_rows if r.get('StartDateMismatch')),
            'endDateMismatch':    sum(1 for r in audit_rows if r.get('EndDateMismatch')),
            'empNoMismatch':      sum(1 for r in audit_rows if r.get('EmpNoMismatch')),
            'nameMismatch':       sum(1 for r in audit_rows if r.get('NameMismatch')),
        },
        'auditRows':    audit_rows,
        'unlinkedEM':   unlinked_em,
        'unlinkedContracts': unlinked_con,
    }
    with open(OUT_JSON, 'w', encoding='utf-8') as f:
        json.dump(json_out, f, ensure_ascii=False, indent=2, default=str)
    safe_print(f"JSON:  {OUT_JSON}")

    # ── Excel output — multiple sheets ──
    with pd.ExcelWriter(OUT_XLSX, engine='openpyxl') as writer:
        # Sheet 1: all audit rows
        df_all = pd.DataFrame(audit_rows)
        df_all.to_excel(writer, sheet_name='LinkAudit', index=False)

        # Sheet 2: matched with flags
        flagged = [r for r in audit_rows if r['FlagCount'] > 0]
        pd.DataFrame(flagged).to_excel(writer, sheet_name='Flagged', index=False)

        # Sheet 3: EM without contract
        em_nc = [r for r in audit_rows if r['LinkStatus'] == 'EmployeeNoContract']
        pd.DataFrame(em_nc).to_excel(writer, sheet_name='EM_NoContract', index=False)

        # Sheet 4: contract without EM
        con_ne = [r for r in audit_rows if r['LinkStatus'] == 'ContractNoEmployee']
        pd.DataFrame(con_ne).to_excel(writer, sheet_name='Contract_NoEmployee', index=False)

        # Sheet 5: unlinked (no valid ID)
        df_unlinked = pd.DataFrame(unlinked_em + unlinked_con)
        if not df_unlinked.empty:
            df_unlinked.to_excel(writer, sheet_name='Unlinked_NoValidID', index=False)
    safe_print(f"Excel: {OUT_XLSX}")

    # ── Report ──
    report = build_report(audit_rows, unlinked_em, unlinked_con,
                          em_records, contract_records)
    OUT_REPORT.write_text(report, encoding='utf-8')
    safe_print(report)
    safe_print(f"\nReport: {OUT_REPORT}")


if __name__ == '__main__':
    main()
